import json, datetime, openpyxl 
from datetime import timedelta
from openpyxl.styles import Alignment, Font, PatternFill
from . models import *
from . tasks import send_transfer_email
from decimal import Decimal
from django.views import View
from django.db.models import Q
from django.db.models import Sum
from django.db import transaction
from django.contrib import messages
from utils.utils import generate_pdf
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse
from finance.models import (
    StockTransaction,
    PurchaseOrderAccount, 
    PurchasesAccount,
    Expense,
    Currency,
    VATTransaction, 
    VATRate,
    Account,
    Cashbook,
    ExpenseCategory,
    AccountBalance,
    AccountTransaction
)
from . utils import (
    calculate_inventory_totals, 
    average_inventory_cost
)
from . forms import (
    BatchForm,
    AddProductForm, 
    addCategoryForm, 
    addTransferForm, 
    DefectiveForm,
    RestockForm, 
    AddDefectiveForm, 
    ServiceForm, 
    AddSupplierForm,
    CreateOrderForm,
    noteStatusForm,
    PurchaseOrderStatus,
    ReorderSettingsForm
)
from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from channels.generic.websocket import  AsyncJsonWebsocketConsumer
from django.shortcuts import render, redirect, get_object_or_404, get_list_or_404
from permissions.permissions import (
    admin_required,
    # sales_required, 
    # accountant_required
)
from utils.account_name_identifier import account_identifier
from loguru import logger


@login_required
def notifications_json(request):
    notifications = StockNotifications.objects.filter(inventory__branch=request.user.branch).values(
        'inventory__product__name', 'type', 'notification', 'inventory__id'
    )
    return JsonResponse(list(notifications), safe=False)

@login_required
def service(request):
    form = ServiceForm()
    if request.method == 'POST':
        form = ServiceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Service successfully created')
            return redirect('inventory:inventory')
        messages.warning(request, 'Invalid form data')
    return redirect('inventory:inventory')
    
@login_required
def batch_code(request):
    if request.method == 'GET':
        
        batch_codes = BatchCode.objects.all().values(
            'id',
            'code'
        )
        logger.info(batch_codes)
        return JsonResponse(list(batch_codes), safe=False)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            code = data.get('batch_code')

            BatchCode.objects.create(code=code)
            return JsonResponse({'success':True}, status=200)
        except Exception as e:
            logger.info(e)
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)

@login_required
def product_list(request): 
    """ for the pos """
    queryset = Inventory.objects.filter(branch=request.user.branch, status=True)
    search_query = request.GET.get('q', '') 
    product_id = request.GET.get('product', '')
    category_id = request.GET.get('category', '')

    if category_id:
        queryset = queryset.filter(product__category__id=category_id)
    if search_query:
        queryset = queryset.filter(
            Q(product__name__icontains=search_query) | 
            Q(product__description__icontains=search_query) 
        )
    if product_id:
        queryset = queryset.filter(id=product_id)

    inventory_data = list(queryset.values(
        'id', 
        'product__id', 
        'product__name', 
        'product__description', 
        'product__category__id', 
        'product__category__name',  
        'product__end_of_day',
        'price', 
        'quantity'
    ))
    
    merged_data = [{
        'inventory_id': item['id'],
        'product_id': item['product__id'],
        'product_name':item['product__name'],
        'description': item['product__description'],
        'category': item['product__category__id'],
        'category_name': item['product__category__name'],
        'end_of_day':item['product__end_of_day'],
        'price': item['price'],
        'quantity': item['quantity'],
    } for item in inventory_data]

    return JsonResponse(merged_data, safe=False)

@login_required
def branches_inventory(request):
    return render(request, 'inventory/branches_inventory.html')

@login_required
def branches_inventory_json(request):
    branches_inventory = Inventory.objects.filter(status=True).values(
        'product__name', 'price', 'quantity', 'branch__name'
    )
    return JsonResponse(list(branches_inventory), safe=False)
        
class AddProductView(LoginRequiredMixin, View):
    form_class = AddProductForm()
    initial = {'key':'value'}
    template_name = 'inventory/add_product.html'
    
    def get(self, request, *args, **kwargs):
        form = self.form_class
        return render(request, self.template_name, {'form': form})
    
    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = AddProductForm(request.POST)
            product_name = request.POST['name']
            
            try:
                product = Product.objects.get(name=product_name)                
                messages.warning(request, 'Product exists ')
                return redirect('inventory:add_product')
            except Product.DoesNotExist:
                if form.is_valid():
                    product = form.save()
                    message = 'Product successfully created'
                    log_action = 'stock in'
                else:
                    messages.warning(request, 'Invalid form data')
                    return redirect('inventory:inventory')
                
            self.create_branch_inventory(product, log_action)
            
            messages.success(request, message)
        return redirect('inventory:inventory')


    def create_branch_inventory(self, product, log_action):
        try:
            inv = Inventory.objects.get(product__name=product.name)
            inv.quantity += product.quantity
            inv.price = product.price
            inv.cost = product.cost
            inv.save()
            self.activity_log(log_action, inv, inv)
        except Inventory.DoesNotExist:
            inventory = Inventory.objects.create(
                product=product,
                quantity=product.quantity,
                price=product.price,
                cost=product.cost,
                branch=self.request.user.branch,
                stock_level_threshold=product.min_stock_level
            )
            self.activity_log(log_action, inventory, inv=0)  
    
    def activity_log(self, action, inventory, inv):
        ActivityLog.objects.create(
            branch = self.request.user.branch,
            user=self.request.user,
            action= action,
            inventory=inventory,
            quantity=inventory.quantity,
            total_quantity=inventory.quantity + inv.quantity if action == 'update' else inventory.quantity
        )

class ProcessTransferCartView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                data = json.loads(request.body)
                branch_name = data['branch_to']
                
                try:
                    branch_to =  Branch.objects.get(name=branch_name)
                except Exception as e:
                    return JsonResponse({'success':False, 'message': f'here {e}'})
                
                transfer = Transfer(
                    transfer_to = branch_to,
                    branch = request.user.branch,
                    user = request.user,
                    transfer_ref = Transfer.generate_transfer_ref(request.user.branch.name, branch_to.name)
                )
                
                for item in data['cart']:
    
                    transfer_item = TransferItems(
                        transfer=transfer,
                        product= Product.objects.get(name=item['product']),
                        price=item['price'],
                        quantity=item['quantity'],
                        from_branch= request.user.branch,
                        to_branch= transfer.transfer_to,
                    )   
                    transfer.save()         
                    transfer_item.save()
                    
                    self.deduct_inventory(transfer_item)
                    self.transfer_update_quantity(transfer_item, transfer)  
                    
                # send email for transfer alert
                transaction.on_commit(lambda: send_transfer_email(request.user.email, transfer.id, transfer.transfer_to.id))
                
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    def deduct_inventory(self, transfer_item):
        logger.info(f'from branch -> {transfer_item.from_branch}')
        branch_inventory = Inventory.objects.get(product__name=transfer_item.product, branch__name=transfer_item.from_branch)
        
        branch_inventory.quantity -= int(transfer_item.quantity)
        branch_inventory.save()
        self.activity_log('Transfer', branch_inventory, transfer_item)
        
    def transfer_update_quantity(self, transfer_item, transfer):
        transfer = Transfer.objects.get(id=transfer.id)
        transfer.quantity += transfer_item.quantity
        transfer.save()
       
    def activity_log(self,  action, inventory, transfer_item,):
        ActivityLog.objects.create(
            invoice = None,
            product_transfer = transfer_item,
            branch = self.request.user.branch,
            user=self.request.user,
            action= action,
            inventory=inventory,
            quantity=transfer_item.quantity,
            total_quantity=inventory.quantity
        )

@login_required
def delete_transfer(request, transfer_id):
    transfer = get_object_or_404(Transfer, id=transfer_id)

    transfer.delete()
    
    return JsonResponse({'success':True})
        
@login_required       
def transfer_details(request, transfer_id):
    transfer = TransferItems.objects.filter(id=transfer_id).values(
        'product__name', 'transfer__transfer_ref', 'quantity', 'price', 'from_branch__name', 'to_branch__name'
    )
    return JsonResponse(list(transfer), safe=False)

@login_required
def inventory(request):
    product_name = request.GET.get('name', '')
    if product_name:
        logger.info(f'{list(Inventory.objects.filter(product__name=product_name, branch=request.user.branch).values())}')
        return JsonResponse(list(Inventory.objects.filter(product__name=product_name, branch=request.user.branch).values()), safe=False)
    return JsonResponse({'error':'product doesnt exists'})


@login_required
def inventory_index(request):
    form = ServiceForm()
    q = request.GET.get('q', '')  
    category = request.GET.get('category', '')    
    
    services = Service.objects.all().order_by('-name')
    inventory = Inventory.objects.filter(branch=request.user.branch, status=True).order_by('product__name')
    
    if category:
        if category == 'inactive':
            inventory = Inventory.objects.filter(branch=request.user.branch, status=False)
        else:
            inventory = inventory.filter(product__category__name=category)
                
    if 'download' and 'excel' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={request.user.branch.name} stock.xlsx'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Add products data
        products = Inventory.objects.all()
        branches = Branch.objects.all().values_list('name', flat=True).distinct()
        row_offset = 0
        for branch in branches:
            worksheet['A' + str(row_offset + 1)] = f'{branch} Products'
            worksheet.merge_cells('A' + str(row_offset + 1) + ':D' + str(row_offset + 1))
            cell = worksheet['A' + str(row_offset + 1)]
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(size=16, bold=True)
            cell.fill = PatternFill(fgColor='AAAAAA', fill_type='solid')

            row_offset += 1 
            
            category_headers = ['Name', 'Cost', 'Price', 'Quantity']
            for col_num, header_title in enumerate(category_headers, start=1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.value = header_title
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            categories = Inventory.objects.filter(branch=request.user.branch).values_list('product__category__name', flat=True).distinct()
            for category in categories:
                products_in_category = products.filter(branch__name=branch, product__category__name=category)
                if products_in_category.exists():
                    worksheet['A' + str(row_offset + 1)] = category
                    cell = worksheet['A' + str(row_offset + 1)]
                    cell.font = Font(color='FFFFFF')
                    cell.fill = PatternFill(fgColor='0066CC', fill_type='solid')
                    worksheet.merge_cells('A' + str(row_offset + 1) + ':D' + str(row_offset + 1))
                    row_offset += 2

                for product in products.filter(branch__name=branch):
                    if category == product.product.category.name:
                        worksheet.append([product.product.name, product.cost, product.price, product.quantity])
                        row_offset += 1

        workbook.save(response)
        return response
    
    all_branches_inventory = Inventory.objects.filter(branch=request.user.branch)
    
    totals = calculate_inventory_totals(all_branches_inventory.filter(status=True))
  
    return render(request, 'inventory/inventory.html', {
        'form': form,
        'services':services,
        'inventory': inventory,
        'search_query': q,
        'category':category,
        'total_price': totals[1],
        'total_cost':totals[0],
    })

@login_required
def edit_service(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    if request.method == 'GET':
        form = ServiceForm(instance=service)
        return render(request, 'inventory/edit_service.html', {'form': form, 'service': service})
    
    elif request.method == 'POST':
        form = ServiceForm(request.POST, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, f'{service.name} successfully edited')
            return redirect('inventory:inventory')
        else:
            messages.warning(request, 'Please correct the errors below')
    
    else:
        messages.warning(request, 'Invalid request')
        return redirect('inventory:inventory')
    
    return render(request, 'inventory/edit_service.html', {'form': form, 'service': service})
        
@login_required   
def inventory_index_json(request):
    inventory = Inventory.objects.filter(branch=request.user.branch, status=True).values(
        'id', 'product__name', 'product__quantity', 'product__id', 'price', 'cost', 'quantity', 'reorder'
    ).order_by('product__name')
    return JsonResponse(list(inventory), safe=False)

@login_required 
@transaction.atomic
def activate_inventory(request, product_id):
    product = get_object_or_404(Inventory, id=product_id)
    product.status=True
    product.save()
    
    ActivityLog.objects.create(
        invoice = None,
        product_transfer = None,
        branch = request.user.branch,
        user=request.user,
        action= 'activated',
        inventory=product,
        quantity=product.quantity,
        total_quantity=product.quantity
    )
    
    messages.success(request, 'Product succefully activated')
    return redirect('inventory:inventory')

@admin_required
@login_required
def edit_inventory(request, product_name):
    inv_product = Inventory.objects.get(product__name=product_name, branch=request.user.branch)

    if request.method == 'POST':
        
        product = Product.objects.get(name=product_name)
        product.name=request.POST['name']
        # product.batch_code=request.POST['batch_code']
        product.description=request.POST['description']
        
        end_of_day = request.POST.get('end_of_day')

        if end_of_day:
            product.end_of_day = True
            
        product.save()
        
        quantity = int(request.POST['quantity'])
        
        # think through
        inv_product.quantity = quantity
             
        inv_product.price = Decimal(request.POST['price'])
        inv_product.cost = Decimal(request.POST['cost'])
        # inv_product.dealer_price = Decimal(request.POST['dealer_price'])
        inv_product.stock_level_threshold = request.POST['min_stock_level']
        inv_product.dealer_price = inv_product.dealer_price
        
        inv_product.save()
        
        ActivityLog.objects.create(
            branch = request.user.branch,
            user=request.user,
            action= 'Edit' if quantity > 0 else 'removed',
            inventory=inv_product,
            quantity=quantity,
            total_quantity=inv_product.quantity,
        )
        
        messages.success(request, f'{product.name} update succesfully')
        return redirect('inventory:inventory')
    return render(request, 'inventory/inventory_form.html', {'product':inv_product, 'title':f'Edit >>> {inv_product.product.name}'})

@login_required
def inventory_detail(request, id):

    inventory = Inventory.objects.get(id=id, branch=request.user.branch)
    logs = ActivityLog.objects.filter(inventory=inventory, branch=request.user.branch)

    sales_data = {}
    stock_in_data = {}
    transfer_data = {}
    labels = []

    for log in logs:
        month_name = log.timestamp.strftime('%B')  
        year = log.timestamp.strftime('%Y')
        month_year = f"{month_name} {year}"  

        if log.action == 'Sale':
            if month_year in sales_data:
                sales_data[month_year] += log.quantity
            else:
                sales_data[month_year] = log.quantity
        elif log.action in ('stock in', 'Update'):
            if month_year in stock_in_data:
                stock_in_data[month_year] += log.quantity
            else:
                stock_in_data[month_year] = log.quantity
        elif log.action == 'Transfer':
            if month_year in transfer_data:
                transfer_data[month_year] += log.quantity
            else:
                transfer_data[month_year] = log.quantity

        if month_year not in labels:
            labels.append(month_year)

    return render(request, 'inventory/inventory_detail.html', {
        'inventory': inventory,
        'logs': logs,
        'sales_data': list(sales_data.values()), 
        'stock_in_data': list(stock_in_data.values()),
        'transfer_data': list(transfer_data.values()),
        'labels': labels
    })


@login_required    
def inventory_transfers(request):
    form = addTransferForm()
    q = request.GET.get('q', '') 
    branch_id = request.GET.get('branch', '')

    transfers = Transfer.objects.filter(branch=request.user.branch).order_by('-time')
    transfer_items = TransferItems.objects.all()
    
    if q:
        transfers = transfers.filter(Q(transfer_ref__icontains=q) | Q(date__icontains=q) )
        
    if branch_id: 
        transfers = transfers.filter(transfer_to__id=branch_id)
        
    if request.method == 'POST':
        form = addTransferForm(request.POST)
        
        if form.is_valid():
            transfer=form.save(commit=False)
            transfer_to = Branch.objects.get(id=int(request.POST['transfer_to']))
            transfer.branch = request.user.branch
            transfer.user = request.user
            transfer.transfer_ref = Transfer.generate_transfer_ref(transfer.branch.name, transfer_to.name)
            
            transfer.save()
            
            return redirect('inventory:add_transfer', transfer.transfer_ref)
            
        messages.success(request, 'Transfer creation failed')
    
    return render(request, 'inventory/transfers.html', {'transfers': transfers,'search_query': q, 'form':form, 'transfer_items':transfer_items })

@login_required
def print_transfer(request, transfer_id):
    
    try:
        transfer = Transfer.objects.get(id=transfer_id)
    except:
        messages.warning(request, 'Transfer doesnt exists')
        return redirect('inventory:transfers')
    
    transfer_items = TransferItems.objects.filter(transfer=transfer)
    
    return render(request, 'inventory/components/ibt.html', {
        'date':datetime.datetime.now(),
        'transfer':transfer, 
        'transfer_items':transfer_items
    })
    
@login_required
@transaction.atomic
def receive_inventory(request):
    transfers =  TransferItems.objects.filter(to_branch=request.user.branch).order_by('-date')
    all_transfers = Transfer.objects.filter(transfer_to=request.user.branch).order_by('-time')    

    if request.method == 'POST':
        transfer_id = request.POST.get('id')  

        try:
            branch_transfer = get_object_or_404(transfers, id=transfer_id)
            transfer_obj = get_object_or_404(all_transfers, id=branch_transfer.transfer.id)
            # validation
            if int(request.POST['quantity']) > branch_transfer.quantity:
                messages.error(request, 'Quantity received cannot be more than quanity transfered') 
                return redirect('inventory:receive_inventory')
                
            if request.POST['received'] == 'true':                                                       
                if int(request.POST['quantity']) != int(branch_transfer.quantity):
                    branch_transfer.over_less_quantity =  branch_transfer.quantity - int(request.POST['quantity']) 
                    branch_transfer.over_less = True
                    branch_transfer.save()
                    
                if Inventory.objects.filter(product=branch_transfer.product, branch=request.user.branch).exists():
                    existing_inventory = Inventory.objects.get(product=branch_transfer.product, branch=request.user.branch)
                    existing_inventory.quantity += int(request.POST['quantity'])
                    existing_inventory.price = branch_transfer.price
                    existing_inventory.dealer_price = Product.objects.get(id=branch_transfer.product.id).dealer_price
                    existing_inventory.save()
                    
                    ActivityLog.objects.create(
                        branch = request.user.branch,
                        user=request.user,
                        action= 'stock in',
                        inventory=existing_inventory,
                        quantity=int(request.POST['quantity']),
                        total_quantity= existing_inventory.quantity,
                        product_transfer=branch_transfer,
                        description = f'received f{request.POST['quantity']} out of {branch_transfer.quantity}' 
                    )
                    messages.success(request, 'Product received')
                    
                else:
                    inventory = Inventory.objects.create(
                        branch=request.user.branch,
                        product=branch_transfer.product,
                        cost=branch_transfer.product.cost,  
                        price=branch_transfer.price,
                        dealer_price = Product.objects.get(id=branch_transfer.product.id).dealer_price,
                        quantity=request.POST['quantity'],
                    )
                    ActivityLog.objects.create(
                        branch = request.user.branch,
                        user=request.user,
                        action= 'stock in',
                        inventory=inventory,
                        quantity=inventory.quantity,
                        total_quantity=inventory.quantity,
                        product_transfer=branch_transfer,
                        description = f'received {request.POST['quantity']} out of {branch_transfer.quantity}'
                    )
                    messages.success(request, 'Product received')
                    
            branch_transfer.quantity_track = branch_transfer.quantity - int(request.POST['quantity'])
            branch_transfer.received_by = request.user
            branch_transfer.received = True
            branch_transfer.description = f'received {request.POST['quantity']} - {branch_transfer.quantity}'
            branch_transfer.save()
            
            transfer_obj.total_quantity_track -= int(request.POST['quantity'])
            transfer_obj.save()
            return redirect('inventory:receive_inventory')  

        except Transfer.DoesNotExist:
            messages.error(request, 'Invalid transfer ID')
            return redirect('inventory:receive_inventory')  
    return render(request, 'inventory/receive_inventory.html', {'r_transfers': transfers, 'transfers':all_transfers})

@login_required
def receive_inventory_json(request):
    transfers =  TransferItems.objects.filter(to_branch=request.user.branch).order_by('-date')
    if request.method ==  'GET':
        transfers = transfers.values(
            'id',
            'date', 
            'quantity',
            'received', 
            'description',
            'date_received',
            'product__name', 
            'from_branch__name',
            'received_by__username'
        )
        return JsonResponse(list(transfers), safe=False)
    
@login_required
@transaction.atomic
def over_less_list_stock(request):
    form = DefectiveForm()
    search_query = request.GET.get('search_query', '')
    
    transfers =  TransferItems.objects.filter(to_branch=request.user.branch)

    if search_query:
        transfers = transfers.filter(
            Q(transfer__product__name__icontains=search_query)|
            Q(transfer__transfer_ref__icontains=search_query)|
            Q(date__icontains=search_query)
        )
        
    def activity_log(action, inventory, branch_transfer):
        ActivityLog.objects.create(
            branch = request.user.branch,
            user=request.user,
            action= action,
            inventory=inventory,
            quantity=branch_transfer.over_less_quantity,
            total_quantity=inventory.quantity,
            product_transfer=branch_transfer
        )
    
    if request.method == 'POST':
        data = json.loads(request.body)
        
        action = data['action']
        transfer_id = data['transfer_id']
        reason = data['reason']
        status = data['status']
        branch_loss = data['branch_loss']
        quantity= data['quantity']
  
        branch_transfer = get_object_or_404(transfers, id=transfer_id)
        transfer = get_object_or_404(Transfer, id=branch_transfer.transfer.id)
        product = Inventory.objects.get(product=branch_transfer.product, branch=request.user.branch)
       
        if int(branch_transfer.over_less_quantity) > 0:
            if action == 'write_off':    
                product.quantity += branch_transfer.over_less_quantity 
                product.save()
                
                description='write_off'
                
                activity_log('Stock in', product, branch_transfer)
                logger.info(f'quantity {quantity}')
                DefectiveProduct.objects.create(
                    product = product,
                    branch = request.user.branch,
                    quantity = quantity,
                    reason = reason,
                    status = status,
                    branch_loss = get_object_or_404(Branch, id=branch_loss),
                )
                
                product.quantity -= branch_transfer.over_less_quantity 
                branch_transfer.over_less_description = description
                branch_transfer.action_by = request.user
                branch_transfer.over_less = False
                
                transfer.defective_status = True
                
                transfer.save()
                branch_transfer.save()
                product.save()
                
                activity_log('write off', product, branch_transfer )
                
                messages.success(request, f'{product.product.name} write-off successfull')        

                return JsonResponse({'success':True}, status=200)
            
            if action == 'accept':
                product.quantity += branch_transfer.over_less_quantity 
                product.save()
                description='returned back'
                activity_log('stock in', product, branch_transfer )
                
                branch_transfer.over_less = False
                branch_transfer.over_less_description = description
                branch_transfer.save()
                
                messages.success(request, f'{product.product.name} accepted back successfully') 

                return JsonResponse({'success':True}, status=200)
            
            if action == 'back':
                description=f'transfered to {branch_transfer.to_branch}'
                product.quantity += branch_transfer.over_less_quantity 
                product.save()
                
                activity_log('stock in', product, branch_transfer )
                
                branch_transfer.received = False
                branch_transfer.quantity = branch_transfer.over_less_quantity
                branch_transfer.save()
                
                product.quantity -= branch_transfer.over_less_quantity 
                product.save()
            
                activity_log('transfer', product, branch_transfer )
                
                branch_transfer.over_less = False
                branch_transfer.over_less_description = description
                branch_transfer.save()
                 
                return JsonResponse({'success':True}, status=200)
            
            return JsonResponse({'success':False, 'messsage':'Invalid form'}, status=400)
            
        return JsonResponse({'success':False, 'messsage':'Invalid form'}, status=400)
                  
    return render(request, 'inventory/over_less_transfers.html', {'over_less_transfers':transfers, 'form':form})

@login_required
@transaction.atomic
def defective_product_list(request):
    form = RestockForm()
    defective_products = DefectiveProduct.objects.filter(branch=request.user.branch)
    
    # loss calculation
    if request.method == 'POST':
        data = json.loads(request.body)
        
        defective_id = data['product_id']
        quantity = data['quantity']
        
        try:
            d_product = DefectiveProduct.objects.get(id=defective_id, branch=request.user.branch)
            product = Inventory.objects.get(product__id=d_product.product.id, branch=request.user.branch)
        except:
            return JsonResponse({'success': False, 'message':'Product doesnt exists'}, status=400)
    
        product.quantity += quantity
        product.status = True if product.status == False else product.status
        product.save()
        
        d_product.quantity -= quantity
        d_product.save()
        
        ActivityLog.objects.create(
            branch = request.user.branch,
            user=request.user,
            action= 'stock in',
            inventory=product,
            quantity=quantity,
            total_quantity=product.quantity,
            description = 'from defective products'
        )
        return JsonResponse({'success':True}, status=200)
    
    quantity = defective_products.aggregate(Sum('quantity'))['quantity__sum'] or 0
    price = defective_products.aggregate(Sum('product__cost'))['product__cost__sum'] or 0
    
    return render(request, 'inventory/defective_products.html', 
        {
            'total_cost': quantity * price,
            'defective_products':defective_products,
            'form':form,
        }
    )
    
@login_required
@transaction.atomic
def create_defective_product(request):
    form = AddDefectiveForm()
    if request.method == 'POST':
        form = AddDefectiveForm(request.POST)

        if form.is_valid():
            branch = request.user.branch
            product = form.cleaned_data['product']
            quantity = form.cleaned_data['quantity']
            
            # validation
            if quantity > product.quantity:
                messages.warning(request, 'Defective quantity cannot more than the products quantity')
                return redirect('inventory:create_defective_product')
            elif quantity == 0:
                messages.warning(request, 'Defective quantity cannot be less than zero')
                return redirect('inventory:create_defective_product')
            
            product.quantity -= quantity
            product.save()
        
            d_obj = form.save(commit=False)
            d_obj.branch = branch
            d_obj.branch_loss = branch
            
            d_obj.save()
            
            ActivityLog.objects.create(
                branch = branch,
                user=request.user,
                action= 'defective',
                inventory=product,
                quantity=quantity,
                total_quantity=product.quantity,
                description = ''
            )
            messages.success(request, 'Product successfuly saved')
            return redirect('inventory:defective_product_list')
        else:
            messages.success(request, 'Invalid form data')
    return render(request, 'inventory/add_defective_product.html', {'form':form})
        

@login_required
def add_inventory_transfer(request):
    form = addTransferForm()
    return render(request, 'inventory/add_transfer.html', {'fornm':form})

@login_required
@admin_required
def delete_inventory(request):
    product_id = request.GET.get('product_id', '')
    
    if product_id:
        inv = Inventory.objects.get(id=product_id, branch=request.user.branch)
        inv.status = False
        inv.save()
    
    ActivityLog.objects.create(
            branch = request.user.branch,
            user=request.user,
            action= 'deactivated',
            inventory=inv,
            quantity=inv.quantity,
            total_quantity=inv.quantity
        )
    
    messages.success(request, 'Product successfully deleted')
    return redirect('inventory:inventory')

@login_required
# @admin_required
def add_product_category(request):
    
    if request.method == 'GET': 
        categories = ProductCategory.objects.all().values()
        logger.info(categories)
        return JsonResponse(list(categories), safe=False)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        category_name = data['name']
        
        if ProductCategory.objects.filter(name=category_name).exists():
            return JsonResponse({'success':False, 'message':'Category Exists'})
        
        category = ProductCategory.objects.create(
            name=category_name
        )
        return JsonResponse({'success':True, 'id': category.id, 'name':category.name})        

@login_required
def reorder_list(request):
    reorder_list = ReorderList.objects.filter(branch=request.user.branch)
        
    if request.method == 'GET':
        if 'download' in request.GET:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={request.user.branch.name} order.xlsx'
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            row_offset = 0
            
            worksheet['A' + str(row_offset + 1)] = f'Order List'
            worksheet.merge_cells('A' + str(row_offset + 1) + ':D' + str(row_offset + 1))
            cell = worksheet['A' + str(row_offset + 1)]
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(size=14, bold=True)
            cell.fill = PatternFill(fgColor='AAAAAA', fill_type='solid')

            row_offset += 1 
                
            category_headers = ['Name', 'Quantity']
            for col_num, header_title in enumerate(category_headers, start=1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.value = header_title
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            categories = reorder_list.values_list('product__product__category__name', flat=True).distinct()
            for category in categories:
                products_in_category = reorder_list.filter(product__product__category__name=category)
                if products_in_category.exists():
                    worksheet['A' + str(row_offset + 1)] = category
                    cell = worksheet['A' + str(row_offset + 1)]
                    cell.font = Font(color='FFFFFF')
                    cell.fill = PatternFill(fgColor='0066CC', fill_type='solid')
                    worksheet.merge_cells('A' + str(row_offset + 1) + ':D' + str(row_offset + 1))
                    row_offset += 2

                for product in reorder_list:
                    if category == product.product.product.category.name:
                        worksheet.append([product.product.product.name])
                        row_offset += 1

            workbook.save(response)
            return response
        return render(request, 'inventory/reorder_list.html', {'form':ReorderSettingsForm()})
        
@login_required
def reorder_list_json(request):
    order_list = ReorderList.objects.filter(branch=request.user.branch).values(
        'id', 
        'product__product__name',  
        'product__quantity',
        'quantity'
    )
    return JsonResponse(list(order_list), safe=False)

@login_required
@transaction.atomic
def clear_reorder_list(request):
    if request.method == 'GET':
        reorders = ReorderList.objects.filter(branch=request.user.branch)
        
        for item in reorders:
            inventory_items = Inventory.objects.filter(id=item.product.id)
            for product in inventory_items:
                product.reorder = False
                product.save()
            
        reorders.delete()
        
        messages.success(request, 'Reoder list success fully cleared')
        return redirect('inventory:reorder_list')

    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data['product_id']
    
        product = ReorderList.objects.get(id=product_id, branch=request.user.branch)
     
        inventory = Inventory.objects.get(id=product.product.id)
    
        product.delete()
        
        inventory.reorder=False
        inventory.save()
        
        return JsonResponse({'success':True}, status=200)
    
        
@login_required
@transaction.atomic
def create_order_list(request):
    if request.method == 'GET':
        products_below_five = Inventory.objects.filter(branch=request.user.branch, quantity__lte = 5).values(
            'id', 
            'product__id', 
            'product__name',
            'product__description',
            'quantity', 
            'reorder',
            'product__category__id',
            'product__category__name'
        )
        return JsonResponse(list(products_below_five), safe=False)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data['id']
        
        product = get_object_or_404(Inventory, id=product_id, branch=request.user.branch)
        ReorderList.objects.create(product=product, branch=request.user.branch)
        
        product.reorder = True
        product.save()
        return JsonResponse({'success': True}, status=201)
    return JsonResponse({'success': False, 'message':'Failed to add the product'}, status=400)
        
    

@login_required
def inventory_pdf(request):
    template_name = 'inventory/reports/inventory_pdf.html'
    category = request.GET.get('category', '')

    inventory = get_list_or_404(Inventory, branch=request.user.branch.id, product__category__name=category) if category else get_list_or_404(Inventory, branch=request.user.branch.id)
    title = f'{category} Inventory' if category else 'All Inventory'
    
    totals = calculate_inventory_totals(inventory)
    
    return generate_pdf(
        template_name, {
            'inventory':inventory,
            'title': f'{request.user.branch.name}: {title}',
            'date':datetime.date.today(),
            'total_cost':totals[0],
            'total_price':totals[1],
            'pdf_name':'Inventory'
        }
    )

@login_required
def transfers_report(request):
    
    template_name = 'inventory/reports/transfers.html'

    view = request.GET.get('view', '')
    choice = request.GET.get('type', '') 
    time_frame = request.GET.get('timeFrame', '')
    branch_id = request.GET.get('branch', '')
    product_id = request.GET.get('product', '')
    transfer_id = request.GET.get('transfer_id', '')
    
    transfers = TransferItems.objects.filter().order_by('-date') 
    
    today = datetime.date.today()
    
    if choice in ['All', '', 'Over/Less']:
        transfers = transfers
    
    if product_id:
        transfers = transfers.filter(product__id=product_id)
    if branch_id:
        transfers = transfers.filter(to_branch_id=branch_id)
    
    def filter_by_date_range(start_date, end_date):
        return transfers.filter(date__range=[start_date, end_date])
    
    date_filters = {
        'All': lambda: transfers, 
        'today': lambda: filter_by_date_range(today, today),
        'yesterday': lambda: filter_by_date_range(today - timedelta(days=1), today - timedelta(days=1)),
        'this week': lambda: filter_by_date_range(today - timedelta(days=today.weekday()), today),
        'this month': lambda: transfers.filter(date__month=today.month, issue_date__year=today.year),
        'this year': lambda: transfers.filter(date__year=today.year),
    }
    
    
    if time_frame in date_filters:
        transfers = date_filters[time_frame]()
         
    if view:
        return JsonResponse(list(transfers.values(
                'date',
                'product__name', 
                'price',
                'quantity', 
                'from_branch__name',
                'from_branch__id',
                'to_branch__id',
                'to_branch__name',
                'received_by__username',
                'date_received',
                'description',
                'received',
                'declined'
            )), 
            safe=False
        )
    
    if transfer_id:
        
        return JsonResponse(list(transfers.filter(id=transfer_id).values(
                'date',
                'product__name', 
                'price',
                'quantity', 
                'from_branch__name',
                'from_branch__id',
                'to_branch__id',
                'to_branch__name',
                'received_by__username',
                'date_received',
                'description',
                'received',
                'declined'
            )), 
            safe=False
        )
    
    return generate_pdf(
        template_name,
        {
            'title': 'Transfers', 
            'date_range': time_frame if time_frame else 'All',
            'report_date': datetime.date.today(),
            'transfers':transfers
        },
    )


@login_required
def reorder_from_notifications(request):
    if request.method == 'GET':
        notifications = StockNotifications.objects.filter(inventory__branch=request.user.branch, inventory__reorder=False, inventory__alert_notification=False).values(
            'quantity',
            'inventory__product__name', 
            'inventory__id', 
            'inventory__quantity' 
        )
        return JsonResponse(list(notifications), safe=False)
    
    if request.method == 'POST':
        # payload
        """
            inventory_id
        """
        
        data = json.loads(request.body)
        
        inventory_id = data['inventory_id']
        action_type = data['action_type']
        
        try:
            inventory = Inventory.objects.get(id=inventory_id)
            stock_notis = StockNotifications.objects.get(inventory=inventory)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
        
        if action_type == 'add':
            inventory.reorder=True
            inventory.save()
            ReorderList.objects.create(
                quantity=0,
                product=inventory, 
                branch=request.user.branch
            )
            
        elif action_type == 'remove':
            inventory.alert_notification=True
            inventory.save()
    
        return JsonResponse({'success':True}, status=200)
    
    return JsonResponse({'success':False, 'message': 'Invalid request'}, status=400)

@login_required
def add_reorder_quantity(request):
    """
    Handles adding a quantity to a reorder item.

    Payload:
    - reorder_id: ID of the reorder item
    - quantity: Quantity to add
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON payload'}, status=400)

    reorder_id = data.get('reorder_id')
    reorder_quantity = data.get('quantity')

    if not reorder_id:
        return JsonResponse({'success': False, 'message': 'Reorder ID is required'}, status=400)

    if not reorder_quantity:
        return JsonResponse({'success': False, 'message': 'Reorder quantity is required'}, status=400)

    try:
        reorder_quantity = int(reorder_quantity)
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Invalid reorder quantity'}, status=400)

    try:
        reorder = ReorderList.objects.get(id=reorder_id)
        reorder.quantity = reorder_quantity
        reorder.save()
        logger.info(reorder.quantity)
        return JsonResponse({'success': True, 'message': 'Reorder quantity updated successfully'}, status=200)
    except ReorderList.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Reorder item not found'}, status=404)
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return JsonResponse({'success': False, 'message': 'An error occurred'}, status=500)

@login_required
def suppliers(request):
    form = AddSupplierForm()
    suppliers = Supplier.objects.all()
    return render(request, 'inventory/suppliers.html', 
        {
            'suppliers':suppliers,
            'form':form
        }
    )

@login_required
def supplier_list_json(request):
    suppliers = Supplier.objects.all().values(
        'id',
        'name'
    )
    return JsonResponse(list(suppliers), safe=False)

@login_required
def create_supplier(request):
    #payload
    """
        name 
        contact
        email
        phone 
        address
    """
    if request.method == 'POST':
        data = json.loads(request.body)
        
        name = data['name']
        contact = data['contact']
        email = data['email']
        phone = data['phone']
        address = data['address']
        
        if not name or not contact or not email or not phone or not address:
            return JsonResponse({'success': False, 'message':'Fill in all the form data'}, status=400)
        
        if Supplier.objects.filter(email=email).exists():
            return JsonResponse({'success': False, 'message':f'Supplier{name} already exists'}, status=400)
        
        supplier = Supplier(
            name = name,
            contact_name = contact,
            email = email,
            phone = phone,
            address = address
        )
        supplier.save()
        logger.info(f'Supplier successfully created {supplier.name}')
        return JsonResponse({'success': True}, status=200)
        
@login_required
def edit_supplier(request, supplier_id):
    # payload
    """
        supplier_id
    """
    
    if request.method == 'POST':
        data = json.loads(request.post)
        supplier_id = data['supplier_id']
        
        if supplier_id:
            try:
                supplier = Supplier.objects.get(id=supplier_id)
            except Exception as e:
                return JsonResponse({'success': False, 'message':f'{supplier_id} does not exists'}, status=400)
                
        form = AddSupplierForm(request.post, instance=supplier)
        
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True}, status=400)
    return JsonResponse({'success': True})

@login_required
def purchase_orders(request):
    form = CreateOrderForm()
    status_form = PurchaseOrderStatus()
    orders = PurchaseOrder.objects.filter(branch = request.user.branch)

    items = PurchaseOrderItem.objects.filter(purchase_order__id=5)

    # Update the 'received' field for each item
    for item in items:
        item.expected_profit
        item.received_quantity
        item.save()
        

    # Perform a bulk update on the 'received' field
    PurchaseOrderItem.objects.bulk_update(items, ['expected_profit', 'received_quantity'])
   
    return render(request, 'inventory/purchase_orders.html', 
        {
            'form':form,
            'orders':orders,
            'status_form':status_form 
        }
    )

    
@login_required
def create_purchase_order(request):
    if request.method == 'GET':
        supplier_form = AddSupplierForm()
        product_form = AddProductForm()
        suppliers = Supplier.objects.all()
        note_form = noteStatusForm()
        batch_form = BatchForm()

        batch_codes = BatchCode.objects.all()
        return render(request, 'inventory/create_purchase_order.html',
            {
                'product_form':product_form,
                'supplier_form':supplier_form,
                'suppliers':suppliers,
                'note_form':note_form,
                'batch_form':batch_form,
                'batch_codes':batch_codes
            }
        )

    if request.method == 'POST':
        try:
            data = json.loads(request.body) 
            purchase_order_data = data.get('purchase_order', {})
            purchase_order_items_data = data.get('po_items', [])
            expenses = data.get('expenses', [])
            cost_allocations = data.get('cost_allocations', [])

            unique_expenses = []

            seen = set()
            for expense in expenses:
                expense_tuple = (expense['name'], expense['amount'])
                if expense_tuple not in seen:
                    seen.add(expense_tuple)
                    unique_expenses.append(expense)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON payload'}, status=400)

        batch = purchase_order_data['batch']
        logger.info(batch)
        supplier_id = purchase_order_data['supplier']
        delivery_date = purchase_order_data['delivery_date']
        status = purchase_order_data['status']
        notes = purchase_order_data['notes']
        total_cost = Decimal(purchase_order_data['total_cost'])
        discount = Decimal(purchase_order_data['discount'])
        tax_amount = Decimal(purchase_order_data['tax_amount'])
        other_amount = Decimal(purchase_order_data['other_amount'])
        payment_method = purchase_order_data.get('payment_method')
    
        if not all([supplier_id, delivery_date, status, total_cost, payment_method]):
            return JsonResponse({'success': False, 'message': 'Missing required fields'}, status=400)

        try:
            supplier = Supplier.objects.get(id=supplier_id)
        except Supplier.DoesNotExist:
            return JsonResponse({'success': False, 'message': f'Supplier with ID {supplier_id} not found'}, status=404)

        try:
            with transaction.atomic():
                purchase_order = PurchaseOrder(
                    order_number=PurchaseOrder.generate_order_number(),
                    batch=batch,
                    supplier=supplier,
                    delivery_date=delivery_date,
                    status=status,
                    notes=notes,
                    total_cost=total_cost,
                    discount=discount,
                    tax_amount=tax_amount,
                    other_amount=other_amount,
                    branch = request.user.branch,
                    is_partial = False,
                    received = False
                )
                purchase_order.save()
                
                purchase_order_items_bulk = []
                for item_data in purchase_order_items_data:
                    product_name = (item_data['product'])
                    quantity = int(item_data['quantity'])
                    unit_cost = Decimal(item_data['price'])
                    actual_unit_cost = Decimal(item_data['actualPrice'])

                    if not all([product_name, quantity, unit_cost]):
                        transaction.set_rollback(True)
                        return JsonResponse({'success': False, 'message': 'Missing fields in item data'}, status=400)

                    try:
                        product = Product.objects.get(name=product_name)
                    except Product.DoesNotExist:
                        transaction.set_rollback(True)
                        return JsonResponse({'success': False, 'message': f'Product with Name {product_name} not found'}, status=404)

                    purchase_order_items_bulk.append(
                        PurchaseOrderItem(
                            purchase_order=purchase_order,
                            product=product,
                            quantity=quantity,
                            unit_cost=unit_cost,
                            actual_unit_cost=actual_unit_cost,
                            received_quantity=0,
                            received=False
                        )
                    )

                    product.batch = product.batch + f'{batch}, '
                    product.price = 0
                    product.save()

                PurchaseOrderItem.objects.bulk_create(purchase_order_items_bulk)

                expense_bulk = []
                for expense in unique_expenses:
                    name = expense['name'] 
                    amount = expense['amount']
                    expense_bulk.append(
                        otherExpenses(
                            purchase_order=purchase_order,
                            name=name,
                            amount=amount
                        )
                    )
                otherExpenses.objects.bulk_create(expense_bulk)

                # cost allocations
                logger.info('processing cost allocations ....')
                costs_list = []
                for cost in cost_allocations:
                    costs_list.append(
                        costAllocationPurchaseOrder(
                            purchase_order = purchase_order,
                            allocated = cost['allocated'],
                            allocationRate = cost['allocationRate'],
                            expense_cost = cost['expCost'],
                            price = cost['price'],
                            quantity = float(cost['price']),
                            product = cost['product'],
                            total = cost['total'],
                            total_buying = cost['totalBuying']
                        )
                    )
                logger.info(f'processing cost allocations .... {costs_list}')
                costAllocationPurchaseOrder.objects.bulk_create(costs_list)

                logger.info('Cost allocations processed :)')
                    
                # update finance accounts (vat, cashbook, expense, account_transaction_log)
                if purchase_order.status in ['Received', 'received']:
                    if_purchase_order_is_received(
                        request, 
                        purchase_order, 
                        tax_amount, 
                        payment_method
                    )       
          
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

        return JsonResponse({'success': True, 'message': 'Purchase order created successfully'})

       
@login_required    
def if_purchase_order_is_received(request, purchase_order, tax_amount, payment_method):
    try:
        currency = Currency.objects.get(default=True)
        rate = VATRate.objects.get(status=True)

        # get account 
        account_details = account_identifier(request, currency, payment_method)
        account_name = account_details['account_name']
        account_type = account_details['account_type']

        account, _ = Account.objects.get_or_create(
            name=account_name,
            type=account_type
        )
        
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,

            defaults={
                'branch':request.user.branch,
                'currency':currency,
                'balance':0
            }
        )

        account_balance.balance -= purchase_order.total_cost
        
        account_balance.save()

        # get or create purchase order category
        category, _ = ExpenseCategory.objects.get_or_create(name='Purchase orders')
        logger.info(category)

        # create an expense and exclude the vat amount
        expense = Expense.objects.create(
            amount = purchase_order.total_cost - purchase_order.tax_amount,
            payment_method = payment_method,
            currency = currency, 
            category = category,
            user = request.user,
            branch = request.user.branch,
            status = False,
            purchase_order = purchase_order,
            description = f'Purchase order: {purchase_order.order_number}',
        )

        # create a cashbook entry
        Cashbook.objects.create(
            expense = expense,
            description = f'Purchase order: {purchase_order.order_number}',
            credit = True,
            amount = purchase_order.total_cost,
            currency = currency,
            branch = request.user.branch
        )

        # create account transaction log
        AccountTransaction.objects.create(
            account = account,
            expense = expense
        )

        # revisit
        # PurchaseOrderAccount.objects.create(
        #     purchase_order = purchase_order,
        #     amount = purchase_order.total_cost - purchase_order.tax_amount,
        #     balance = 0,
        #     expensed = False
        # )

        # create a vat entry
        VATTransaction.objects.create(
            purchase_order = purchase_order,
            vat_type='Input',
            vat_rate = rate.rate,
            tax_amount = tax_amount
        )

        logger.info('done')
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success':False, 'message':f'currency doesnt exists'})
    except VATRate.DoesNotExist:
        return JsonResponse({'success':False, 'message':f'Make sure you have a stipulated vat rate in the system'})

@login_required
def delete_purchase_order(request, purchase_order_id):
    try:
        # Retrieve the purchase order
        purchase_order = PurchaseOrder.objects.get(id=purchase_order_id)

        if purchase_order.received:
            return JsonResponse({'success': False, 'message': 'Cannot delete a received purchase order'}, status=400)

        with transaction.atomic():
            
            # Reverse related account transactions
            currency = Currency.objects.get(default=True)

            account_transaction = AccountTransaction.objects.filter(expense__purchase_order=purchase_order).first()
            if account_transaction:
                account_balance = AccountBalance.objects.get(account=account_transaction.account)

                # Reverse account balance adjustments
                account_balance.balance += purchase_order.total_cost
                account_balance.save()

                # Delete account transaction log
                account_transaction.delete()

            # Reverse Cashbook entry
            cashbook_entry = Cashbook.objects.filter(expense__purchase_order=purchase_order).first()
            if cashbook_entry:
                cashbook_entry.delete()

            # Reverse the VAT transaction
            vat_transaction = VATTransaction.objects.filter(purchase_order=purchase_order).first()
            if vat_transaction:
                vat_transaction.delete()

            # Reverse the Expense record
            expense = Expense.objects.filter(purchase_order=purchase_order).first()
            if expense:
                expense.delete()

            # Reverse other expenses related to the purchase order
            other_expenses = otherExpenses.objects.filter(purchase_order=purchase_order)
            if other_expenses.exists():
                other_expenses.delete()

            # Remove PurchaseOrderItems
            PurchaseOrderItem.objects.filter(purchase_order=purchase_order).delete()

            # Finally, delete the purchase order itself
            purchase_order.delete()

        return JsonResponse({'success': True, 'message': 'Purchase order deleted successfully'})
    except PurchaseOrder.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Purchase order not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
@transaction.atomic
def change_purchase_order_status(request, order_id):
    try:
        purchase_order = PurchaseOrder.objects.get(id=order_id)
    except PurchaseOrder.DoesNotExist:
        return JsonResponse({'error': f'Purchase order with ID: {order_id} does not exist'}, status=404)

    try:
        data = json.loads(request.body)
        status = data['status']
        
        if status:
            purchase_order.status=status

            with transaction.atomic():
                if purchase_order.status == 'received':
                    purchase_order.save()

                    tax_amount = purchase_order.tax_amount
                    payment_method = purchase_order.payment_method

                    if_purchase_order_is_received(
                        request, 
                        purchase_order, 
                        tax_amount,
                        payment_method
                    )
            
            return JsonResponse({'success':True}, status=200)
        else:
            return JsonResponse({'success':False, 'message':'Status is required'}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON payload'}, status=400)

@login_required
def print_purchase_order(request, order_id):
    try:
        purchase_order = PurchaseOrder.objects.get(id=order_id)
    except PurchaseOrder.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
        return redirect('inventory:purchase_orders')
    
    try:
        purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)
    except PurchaseOrderItem.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
        return redirect('inventory:purchase_orders')
    
    return render(request, 'inventory/print_purchase_order.html', 
        {
            'orders':purchase_order_items,
            'purchase_order':purchase_order
        }
    )

@login_required
def purchase_order_detail(request, order_id):
    try:
        purchase_order = PurchaseOrder.objects.get(id=order_id)
    except PurchaseOrder.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
        return redirect('inventory:purchase_orders')
    
    try:
        purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)
    except PurchaseOrderItem.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
        return redirect('inventory:purchase_orders')

    items = costAllocationPurchaseOrder.objects.filter(purchase_order__id=order_id)
    expenses = otherExpenses.objects.filter(purchase_order__id=order_id)
    total_expense_sum = expenses.aggregate(total_expense=Sum('amount'))['total_expense'] or 0

    logger.info(items)
    
    return render(request, 'inventory/purchase_order_detail.html', 
        {
            'items':items,
            'expenses':expenses,
            'orders':purchase_order_items,
            'purchase_order':purchase_order,
            'total_expenses':total_expense_sum 

        }
    )
    
@login_required
def delete_purchase_order(request, purchase_order_id):
    if request.method != "DELETE":
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

    try:
        purchase_order = PurchaseOrder.objects.get(id=purchase_order_id)
    except PurchaseOrder.DoesNotExist:
        return JsonResponse({'success': False, 'message': f'Purchase order with ID {purchase_order_id} not found'}, status=404)

    try:
        purchase_order.delete()
        return JsonResponse({'success': True, 'message': 'Purchase order deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
def receive_order(request, order_id):
    try:
        purchase_order = PurchaseOrder.objects.get(id=order_id)
    except PurchaseOrder.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
        return redirect('inventory:purchase_orders')
    
    try:
        purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)
    except PurchaseOrderItem.DoesNotExist:
        messages.warning(request, f'Purchase order with ID: {order_id} does not exists')
        return redirect('inventory:purchase_orders')
    
    
    return render(request, 'inventory/receive_order.html', 
        {
            'orders':purchase_order_items,
            'purchase_order':purchase_order
        }
    )

@login_required
@transaction.atomic
def process_received_order(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order_item_id = data.get('id')
            quantity = data.get('quantity', 0)
            selling_price = data.get('selling_price', 0)
            dealer_price = data.get('dealer_price', 0)
            expected_profit = data.get('expected_profit', 0)

            logger.info(data)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON payload'}, status=400)

        if quantity == 0:
            return JsonResponse({'success': False, 'message': 'Quantity cannot be zero.'}, status=400)

        try:
            order_item = PurchaseOrderItem.objects.get(id=order_item_id)
        except PurchaseOrderItem.DoesNotExist:
            return JsonResponse({'success': False, 'message': f'Purchase Order Item with ID: {order_item_id} does not exist'}, status=404)

        # Update expected profit and check quantity
        cost = order_item.actual_unit_cost
        
        #if quantity > order_item.quantity:
        #    return JsonResponse({'success': False, 'message': 'Quantity cannot be more than ordered quantity.'})

        # Update the order item with received quantity
        order_item.receive_items(quantity)
        order_item.expected_profit = expected_profit
        #order_item.check_received() revisit the model method

        # Update or create inventory
        try:
            product = Product.objects.get(id=order_item.product.id)
            product.quantity = quantity
            product.price = selling_price
            product.dealer_price = dealer_price
            product.save()
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'message': f'Product with ID: {order_item.product.id} does not exist'}, status=404)

        try:
            inventory = Inventory.objects.get(product=product, branch=request.user.branch)
            # Update existing inventory
            inventory.cost = cost
            inventory.price = selling_price
            inventory.dealer_price = dealer_price
            inventory.quantity += quantity
            inventory.batch += f'{order_item.batch}, ',
            inventory.save()
        except Inventory.DoesNotExist:
            # Create a new inventory object if it does not exist
            inventory = Inventory(
                product=product,
                branch=request.user.branch,
                cost=cost,
                price=selling_price,
                dealer_price=dealer_price,
                quantity=quantity,
                stock_level_threshold=product.min_stock_level,
                reorder=False,
                alert_notification=True,
                batch = f'{order_item.batch}, '
            )
            inventory.save()

        # Prepare activity log for this transaction
        log = ActivityLog(
            purchase_order=order_item.purchase_order,
            branch=request.user.branch,
            user=request.user,
            action='stock in',
            inventory=inventory,
            quantity=quantity,
            description=f'Stock in from {order_item.purchase_order.order_number}',
            total_quantity=inventory.quantity
        )
        log.save()

        # Save updated order item
        order_item.save()

        return JsonResponse({'success': True, 'message': 'Inventory updated successfully'}, status=200)

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

@login_required
def mark_purchase_order_done(request, po_id):
    purchase_order = PurchaseOrder.objects.get(id=po_id)
    purchase_order.received = True
    purchase_order.save()

    messages.info(request, 'Purchase order done')
    return redirect('inventory:purchase_orders')

@login_required
def edit_purchase_order(request, po_id):
    if request.method == 'GET':
        purchase_order = PurchaseOrder.objects.get(id=po_id)
        supplier_form = AddSupplierForm()
        product_form = AddProductForm()
        suppliers = Supplier.objects.all()
        note_form = noteStatusForm()
        batch_form = BatchForm()

        batch_codes = BatchCode.objects.all()

        return render(request, 'inventory/edit_purchase_order.html', {
            'purchase_order':purchase_order,
            'product_form':product_form,
            'supplier_form':supplier_form,
            'suppliers':suppliers,
            'note_form':note_form,
            'batch_form':batch_form,
            'batch_codes':batch_codes

         })
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            purchase_order_data = data.get('purchase_order', {})
            purchase_order_items_data = data.get('po_items', [])
            expenses = data.get('expenses', [])
            cost_allocations = data.get('cost_allocations', [])

            # remove duplicates
            unique_expenses = []
            seen = set()
            for expense in expenses:
                expense_tuple = (expense['name'], expense['amount'])
                if expense_tuple not in seen:
                    seen.add(expense_tuple)
                    unique_expenses.append(expense)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON payload'}, status=400)

        batch = purchase_order_data['batch']
        supplier_id = purchase_order_data['supplier']
        delivery_date = purchase_order_data['delivery_date']
        status = purchase_order_data['status']
        notes = purchase_order_data['notes']
        total_cost = Decimal(purchase_order_data['total_cost'])
        discount = Decimal(purchase_order_data['discount'])
        tax_amount = Decimal(purchase_order_data['tax_amount'])
        other_amount = Decimal(purchase_order_data['other_amount'])
        payment_method = purchase_order_data.get('payment_method')
    
        if not all([delivery_date, status, total_cost, payment_method]):
            return JsonResponse({'success': False, 'message': 'Missing required fields'}, status=400)

        try:
            supplier = Supplier.objects.get(id=1)
        except Supplier.DoesNotExist:
            return JsonResponse({'success': False, 'message': f'Supplier with ID {supplier_id} not found'}, status=404)

        try:
            with transaction.atomic():
                purchase_order = PurchaseOrder(
                    batch = batch,
                    order_number=PurchaseOrder.generate_order_number(),
                    supplier=supplier,
                    delivery_date=delivery_date,
                    status=status,
                    notes=notes,
                    total_cost=total_cost,
                    discount=discount,
                    tax_amount=tax_amount,
                    other_amount=other_amount,
                    branch = request.user.branch,
                    is_partial = False,
                    received = False
                )
                purchase_order.save()
                
                purchase_order_items_bulk = []
                for item_data in purchase_order_items_data:
                    product_name = (item_data['product'])
                    quantity = int(item_data['quantity'])
                    unit_cost = Decimal(item_data['price'])
                    actual_unit_cost = Decimal(item_data['actualPrice'])

                    logger.info(f'quantity: {quantity}')

                    if not all([product_name, quantity, unit_cost]):
                        transaction.set_rollback(True)
                        return JsonResponse({'success': False, 'message': 'Missing fields in item data'}, status=400)

                    try:
                        product = Product.objects.get(name=product_name)
                    except Product.DoesNotExist:
                        transaction.set_rollback(True)
                        return JsonResponse({'success': False, 'message': f'Product with Name {product_name} not found'}, status=404)

                    purchase_order_items_bulk.append(
                        PurchaseOrderItem(
                            purchase_order=purchase_order,
                            product=product,
                            quantity=quantity,
                            unit_cost=unit_cost,
                            actual_unit_cost=actual_unit_cost,
                            received_quantity=0,
                            received=False
                        )
                    )

                    product.price = 0
                    product.save()

                PurchaseOrderItem.objects.bulk_create(purchase_order_items_bulk)

                expense_bulk = []
                for expense in unique_expenses:
                    name = expense['name'] 
                    amount = expense['amount']
                    expense_bulk.append(
                        otherExpenses(
                            purchase_order=purchase_order,
                            name=name,
                            amount=amount
                        )
                    )
                otherExpenses.objects.bulk_create(expense_bulk)

                logger.info('processing cost allocations ....')
                costs_list = []
                for cost in cost_allocations:
                    costs_list.append(
                        costAllocationPurchaseOrder(
                            purchase_order = purchase_order,
                            allocated = cost['allocated'],
                            allocationRate = cost['allocationRate'],
                            expense_cost = cost['expCost'],
                            price = cost['price'],
                            quantity = int(cost['quantity']),
                            product = cost['product'],
                            total = cost['total'],
                            total_buying = cost['totalBuying']
                        )
                    )
                logger.info(f'processing cost allocations .... {costs_list}')
                costAllocationPurchaseOrder.objects.bulk_create(costs_list)
                    
                # update finance accounts (vat, cashbook, expense, account_transaction_log)
                if purchase_order.status in ['Received', 'received']:
                    if_purchase_order_is_received(
                        request, 
                        purchase_order, 
                        tax_amount, 
                        payment_method
                    )
                
                remove_purchase_order(po_id, request)
                
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

        return JsonResponse({'success': True, 'message': 'Purchase order created successfully'})

def remove_purchase_order(purchase_order_id, request):
    try:
        # Retrieve the purchase order
        purchase_order = PurchaseOrder.objects.get(id=purchase_order_id)

        logger.info(f'{purchase_order} remove')

        with transaction.atomic():
            
            # Reverse related account transactions
            currency = Currency.objects.filter(default=True).first()

            account_transaction = AccountTransaction.objects.filter(expense__purchase_order=purchase_order).first()
            if account_transaction:
                account_balance = AccountBalance.objects.get(account=account_transaction.account)

                # Reverse account balance adjustments
                account_balance.balance += purchase_order.total_cost
                account_balance.save()

                # Delete account transaction log
                account_transaction.delete()

            # Reverse Cashbook entry
            cashbook_entry = Cashbook.objects.filter(expense__purchase_order=purchase_order).first()
            if cashbook_entry:
                cashbook_entry.delete()

            # Reverse the VAT transaction
            vat_transaction = VATTransaction.objects.filter(purchase_order=purchase_order).first()
            if vat_transaction:
                vat_transaction.delete()

            # Reverse the Expense record
            expense = Expense.objects.filter(purchase_order=purchase_order).first()
            if expense:
                expense.delete()

            # Reverse other expenses related to the purchase order
            other_expenses = otherExpenses.objects.filter(purchase_order=purchase_order)
            if other_expenses.exists():
                other_expenses.delete()

            #deduct product quantity
            items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)
            products = Inventory.objects.all()

            for item in items:
                for prod in products:
                    if item.product == prod.product:
                        prod.quantity -= item.received_quantity
                        prod.save()
                        
                        ActivityLog.objects.create(
                            branch = request.user.branch,
                            user= request.user,
                            action= 'delete',
                            inventory=prod,
                            quantity=item.received_quantity,
                            total_quantity=prod.quantity
                        )

            # Remove PurchaseOrderItems
            PurchaseOrderItem.objects.filter(purchase_order=purchase_order).delete()

            # Finally, delete the purchase order itself
            purchase_order.delete()
    except Exception as e:
        logger.info(e)

@login_required
def edit_purchase_order_data(request, po_id):
    try:
        expenses = otherExpenses.objects.filter(purchase_order__id=po_id).values()
    
        purchase_order_items = PurchaseOrderItem.objects.filter(purchase_order__id=po_id).values(
            'purchase_order__id',
            'product__name',
            'quantity',
            'unit_cost',
            'actual_unit_cost',
            'expected_profit'
        )

        return JsonResponse({'success':True, 'po_items':list(purchase_order_items), 'expenses':list(expenses)})

    except Exception as e:
        return JsonResponse({"success":False, 'message':f'{e}'})
    
@login_required
def product(request):

    if request.method == 'POST':
        # payload
        """
            name,
            cost: float,
            quantity: int,
            category,
            tax_type,
            min_stock_level,
            description
        """
        try:
            data = json.loads(request.body)
            
        except Exception as e:
            return JsonResponse({'success':False, 'message':'Invalid data'})

        # validation for existance
        if Product.objects.filter(name=data['name']).exists():
            return JsonResponse({'success':False, 'message':f'Product {data['name']} exists'})

        try:
            category = ProductCategory.objects.get(id=data['category'])
        except ProductCategory.DoesNotExist:
            return JsonResponse({'success':False, 'message':f'Category Doesnt Exists'})
        
        logger.info(category)
        
        product = Product.objects.create(
            batch = '',
            name = data['name'],
            price = 0,
            cost = 0,
            quantity = 0,
            category = category,
            tax_type = data['tax_type'],
            min_stock_level = data['min_stock_level'],
            description = data['description'], 
            end_of_day = True if data['end_of_day'] else False,
            service = True if data['service'] else False,
        )
        product.save()
        
        Inventory.objects.create(
            product = product,
            branch = request.user.branch, # To be specific
            cost = product.cost,
            price = product.price,
            dealer_price = 0,
            quantity = product.quantity
        )
        return JsonResponse({'success':True})
            
    if request.method == 'GET':
        products = Product.objects.all().values(
            'id',
            'name',
        )
        return JsonResponse(list(products), safe=False)
    
    return JsonResponse({'success':False, 'message':'Invalid request'})

@login_required
def reorder_settings(request):
    """ method to set reorder settings"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            quantity_suggestion = data.get('suggestion')
            order_enough = data.get('enough')
            supplier = data.get('supplier', 'all')
            days_from = data.get('from')
            days_to = data.get('to')

            if not quantity_suggestion or not order_enough or not supplier:
                return JsonResponse({'success':False, 'message':'Please fill all required data.'})

            settings = reorderSettings.objects.get_or_create(id=1)

            settings.supplier=supplier,
            settings.quantity_suggestion = True if quantity_suggestion else False,
            settings.order_enough_stock = True if order_enough else False
            
            if order_enough:
                # validate days 
                if not days_from or not days_to:
                    return JsonResponse({'success':False, 'message':'Please fill in the days.'})
                settings.number_of_days_from = days_from
                settings.number_of_days_to = days_to
                settings.save()
            
            return JsonResponse({'success':True, 'message':'Reorder Settings Succefully Saved.'}, status=200)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
    
    if request.method == 'GET':
        try:
            settings = reorderSettings.objects.filter(id=1).values()
            JsonResponse({'success':True, 'data':settings}, status=200)
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}'}, status=400)
        return JsonResponse({'success':False, 'message':'Invalid request'}, status=500)

        
                    
        