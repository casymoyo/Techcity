import csv
from .models import *
from decimal import Decimal
from io import BytesIO
from users.models import User
from company.models import Branch
from .consumers import CashTransferConsumer 
from xhtml2pdf import pisa 
from django.views import View
from django.db.models import Q
from twilio.rest import Client
from datetime import timedelta
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from django.contrib import messages
from utils.utils import generate_pdf
from django.http import JsonResponse
from utils.utils import generate_pdf
from asgiref.sync import async_to_sync, sync_to_async
from inventory.models import Inventory
from channels.layers import get_channel_layer
import json, datetime, os, boto3, openpyxl 
from utils.account_name_identifier import account_identifier
from .tasks import send_invoice_email_task, send_account_statement_email
from pytz import timezone as pytz_timezone 
from openpyxl.styles import Alignment, Font
from . utils import calculate_expenses_totals
from django.utils.dateparse import parse_date
from django.templatetags.static import static
from django.db.models import Sum, DecimalField
from inventory.models import ActivityLog, Product
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.core.mail import send_mail, EmailMessage
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from . forms import (
    ExpenseForm, 
    ExpenseCategoryForm, 
    CurrencyForm, 
    InvoiceForm, 
    CustomerForm, 
    TransferForm, 
    CashWithdrawForm, 
    cashWithdrawExpenseForm,
    customerDepositsForm,
    customerDepositsRefundForm
)
from django.contrib.auth import authenticate
from loguru import logger
from .tasks import send_expense_creation_notification

def get_previous_month():
    first_day_of_current_month = datetime.datetime.now().replace(day=1)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    return last_day_of_previous_month.month

def get_current_month():
    return datetime.datetime.now().month

class Finance(View):
    # authentication loginmixin
    template_name = 'finance/finance.html'

    def get(self, request, *args, **kwargs):
        
        balances = AccountBalance.objects.filter(branch=request.user.branch)
    
        recent_sales = Sale.objects.filter(transaction__branch=request.user.branch).order_by('-date')[:5]

        expenses_by_category = Expense.objects.values('category__name').annotate(
            total_amount=Sum('amount', output_field=DecimalField())
        )
        
        context = {
            'balances': balances,
            'recent_transactions': recent_sales,
            'expenses_by_category': expenses_by_category,
        }
        return render(request, self.template_name, context)

@login_required
def expenses(request):
    form = ExpenseForm()
    cat_form = ExpenseCategoryForm()

    if request.method == 'GET':
        filter_option = request.GET.get('filter', 'today')
        download = request.GET.get('download')
        
        now = datetime.datetime.now()
        end_date = now
        
        if filter_option == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_week':
            start_date = now - timedelta(days=now.weekday())
        elif filter_option == 'yesterday':
            start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == 'this_month':
            start_date = now.replace(day=1)
        elif filter_option == 'last_month':
            start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
        elif filter_option == 'this_year':
            start_date = now.replace(month=1, day=1)
        elif filter_option == 'custom':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        else:
            start_date = now - timedelta(days=now.weekday())
            end_date = now
            
        expenses = Expense.objects.filter(issue_date__gte=start_date, issue_date__lte=end_date).order_by('issue_date')
        
        if download:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="expenses_report_{filter_option}.csv"'

            writer = csv.writer(response)
            writer.writerow(['Date', 'Description', 'Done By', 'Amount'])

            total_expense = 0  
            for expense in expenses:
                total_expense += expense.amount

                writer.writerow([
                    expense.issue_date,
                    expense.description,
                    expense.user.first_name,
                    expense.amount,
                ])

            writer.writerow(['Total', '', '', total_expense])
            
            return response
        
        return render(request, 'finance/expenses.html', 
            {
                'form':form,
                'cat_form':cat_form,
                'expenses':expenses,
                'filter_option': filter_option,
            }
        )
    if request.method == 'POST':
        #payload
        """
            {
                amount:float
                description:str
                category:id (int)
                payment_method:str
                currency:id
            }
        """
    try:
        data = json.loads(request.body)
        
        amount = data.get('amount')
        description = data.get('description')
        category = data.get('category')
        payment_method = data.get('payment_method')
        currency_id = data.get('currency')

        logger.info(currency_id)
        
        if not amount or not description or not category:
            return JsonResponse({'success':False, 'message':'Missing fields: amount, description, category.'})
        
        try:
            category = ExpenseCategory.objects.get(id=category)
        except ExpenseCategory.DoesNotExist:
            return JsonResponse({'success':False, 'message':f'Category with ID: {category}, doesn\'t exists.'})
        
        currency = get_object_or_404(Currency, id=currency_id)

        account_details = account_identifier(request, currency, payment_method)
        account_name = account_details['account_name']
        account_type = account_details['account_type']

        account, _ = Account.objects.get_or_create(
            account_name,
            type=account_type
        )
        account_balance, _ = AccountBalance.objects.get_or_create(account=account)

        logger.info(f'Account: {account}')
        logger.info(f'Account Balance: {account_balance}')

        if account_balance.balance < Decimal(amount):
            return JsonResponse({'success':False, 'message':f'{account_name} have insufficient balance.'})
        
        account_balance.balance -= Decimal(amount)
        account_balance.save()
        
        expense = Expense.objects.create(
            amount = amount,
            category = category,
            user = request.user,
            currency = currency,
            payment_method = payment_method,
            description = description,
            branch = request.user.branch
        )
        
        Cashbook.objects.create(
            amount = amount,
            expense = expense,
            currency = currency,
            credit = True,
            description=f'Expense ({expense.description[:20]})',
            branch = request.user.branch
        )

        send_expense_creation_notification(expense.id)
        
        return JsonResponse({'success': True, 'messages':'Expense successfully created'}, status=201)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required   
def get_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)
    data = {
        'id': expense.id,
        'amount': expense.amount,
        'description': expense.description,
        'category': expense.category.id
    }
    return JsonResponse({'success': True, 'data': data})

@login_required      
def add_or_edit_expense(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            amount = data.get('amount')
            description = data.get('description')
            category_id = data.get('category')
            expense_id = data.get('id')

            if not amount or not description or not category_id:
                return JsonResponse({'success': False, 'message': 'Missing fields: amount, description, category.'})
            
            category = get_object_or_404(ExpenseCategory, id=category_id)

            if expense_id:  
                expense = get_object_or_404(Expense, id=expense_id)
                before_amount = expense.amount
                
                expense.amount = amount
                expense.description = description
                expense.category = category
                expense.save()
                message = 'Expense successfully updated'
                
                try:
                    cashbook_expense = Cashbook.objects.get(expense=expense)
                    expense_amount = Decimal(expense.amount)
                    if cashbook_expense.amount < expense_amount:
                        cashbook_expense.amount = expense_amount
                        cashbook_expense.description = cashbook_expense.description + f'Expense (update from {before_amount} to {cashbook_expense.amount})'
                    else:
                        cashbook_expense.amount -= cashbook_expense.amount - expense_amount
                        cashbook_expense.description = cashbook_expense.description + f'(update from {before_amount} to {cashbook_expense.amount})'
                    cashbook_expense.save()
                except Exception as e:
                    return JsonResponse({'success': False, 'message': str(e)}, status=400)
            return JsonResponse({'success': True, 'message': message}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

@login_required
def add_expense_category(request):
    categories = ExpenseCategory.objects.all().values()
    
    if request.method == 'POST':
        data = json.loads(request.body)
        category = data['name']
        logger.info(data)
        
        if ExpenseCategory.objects.filter(name=category).exists():
            return JsonResponse({'success':False, 'message':f'Category with ID {category} Exists.'}, status=400)
        
        ExpenseCategory.objects.create(
            name=category
        )
        return JsonResponse({'success':True}, status=201)
    return JsonResponse(list(categories), safe=False)

@login_required
@transaction.atomic
def delete_expense(request, expense_id):
    if request.method == 'DELETE':
        try:
            expense = get_object_or_404(Expense, id=expense_id)
            expense.cancel = True
            expense.save()
            
            Cashbook.objects.create(
                amount=expense.amount,
                debit=True,
                credit=False,
                description=f'Expense ({expense.description}): cancelled'
            )
            return JsonResponse({'success': True, 'message': 'Expense successfully deleted'})
        except Exception as e:
             return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

@login_required
def update_expense_status(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            expense_id = data.get('id')
            status = data.get('status')

            expense = Expense.objects.get(id=expense_id)
            expense.status = status
            expense.save()

            return JsonResponse({'success': True, 'message': 'Status updated successfully.'})
        except Expense.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Expense not found.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@login_required
def invoice(request):
    form = InvoiceForm()
    invoices = Invoice.objects.filter(branch=request.user.branch, status=True).order_by('-invoice_number')

    query_params = request.GET
    if query_params.get('q'):
        search_query = query_params['q']
        invoices = invoices.filter(
            Q(customer__name__icontains=search_query) |
            Q(invoice_number__icontains=search_query) |
            Q(issue_date__icontains=search_query)
        )

    user_timezone_str = request.user.timezone if hasattr(request.user, 'timezone') else 'UTC'
    user_timezone = pytz_timezone(user_timezone_str)  

    def filter_by_date_range(start_date, end_date):
        start_datetime = user_timezone.localize(
            timezone.datetime.combine(start_date, timezone.datetime.min.time())
        )
        end_datetime = user_timezone.localize(
            timezone.datetime.combine(end_date, timezone.datetime.max.time())
        )
        return invoices.filter(issue_date__range=[start_datetime, end_datetime])

    now = timezone.now().astimezone(user_timezone)
    today = now.date()

    now = timezone.now() 
    today = now.date()  
    
    date_filters = {
        'today': lambda: filter_by_date_range(today, today),
        'yesterday': lambda: filter_by_date_range(today - timedelta(days=1), today - timedelta(days=1)),
        't_week': lambda: filter_by_date_range(today - timedelta(days=today.weekday()), today),
        'l_week': lambda: filter_by_date_range(today - timedelta(days=today.weekday() + 7), today - timedelta(days=today.weekday() + 1)),
        't_month': lambda: invoices.filter(issue_date__month=today.month, issue_date__year=today.year),
        'l_month': lambda: invoices.filter(issue_date__month=today.month - 1 if today.month > 1 else 12, issue_date__year=today.year if today.month > 1 else today.year - 1),
        't_year': lambda: invoices.filter(issue_date__year=today.year),
    }

    if query_params.get('day') in date_filters:
        invoices = date_filters[query_params['day']]()

    total_partial = invoices.filter(payment_status='Partial').aggregate(Sum('amount'))['amount__sum'] or 0
    total_paid = invoices.filter(payment_status='Paid').aggregate(Sum('amount'))['amount__sum'] or 0
    total_amount = invoices.aggregate(Sum('amount'))['amount__sum'] or 0

    return render(request, 'finance/invoices/invoice.html', {
        'form': form,
        'invoices': invoices,
        'total_paid': total_paid,
        'total_due': total_partial,
        'total_amount': total_amount,
    })


@login_required
@transaction.atomic 
def update_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    customer_account = get_object_or_404(CustomerAccount, customer=invoice.customer)
    customer_account_balance = get_object_or_404(
        CustomerAccountBalances, account=customer_account, currency=invoice.currency
    )

    if request.method == 'POST':
        data = json.loads(request.body)
        amount_paid = Decimal(data['amount_paid'])

        invoice = Invoice.objects.select_for_update().get(pk=invoice.pk)
        customer_account_balance = CustomerAccountBalances.objects.select_for_update().get(pk=customer_account_balance.pk)

        if amount_paid <= 0:
            return JsonResponse({'success': False, 'message': 'Invalid amount paid.'}, status=400)

        if amount_paid >= invoice.amount_due:
            invoice.payment_status = Invoice.PaymentStatus.PAID
            invoice.amount_due = 0
        else:
            invoice.amount_due -= amount_paid

        invoice.amount_paid += amount_paid
        
        # get the latest payment for the invoice
        latest_payment = Payment.objects.filter(invoice=invoice).order_by('-payment_date').first()
        amount_due = latest_payment.amount_due - amount_paid 

        payment = Payment.objects.create(
            invoice=invoice,
            amount_paid=amount_paid,
            amount_due=amount_due, 
            payment_method=data['payment_method'],
            user=request.user
        )

        account, _ = Account.objects.get_or_create(
            name=f"{request.user.branch} {invoice.currency.name} {payment.payment_method.capitalize()} Account",
            type=Account.AccountType[payment.payment_method.upper()] 
        )
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,
            currency=invoice.currency,
            branch=request.user.branch,
            defaults={'balance': 0}
        )

        account_balance.balance += amount_paid
        if customer_account_balance.balance < 0:
            customer_account_balance.balance += amount_paid
        else:
            customer_account_balance.balance -= amount_paid

        account_balance.save()
        customer_account_balance.save()
        invoice.save()
        payment.save()

        return JsonResponse({'success': True, 'message': 'Invoice successfully updated'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}) 


@login_required
@transaction.atomic 
def create_invoice(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            invoice_data = data['data'][0]  
            items_data = data['items']
           
            # get currency
            currency = Currency.objects.get(id=invoice_data['currency'])
            
            # create or get accounts
            account_types = {
                'cash': Account.AccountType.CASH,
                'bank': Account.AccountType.BANK,
                'ecocash': Account.AccountType.ECOCASH,
            }

            account_name = f"{request.user.branch} {currency.name} {invoice_data['payment_method'].capitalize()} Account"
            
            account, _ = Account.objects.get_or_create(name=account_name, type=account_types[invoice_data['payment_method']])
            
            account_balance, _ = AccountBalance.objects.get_or_create(
                account=account,
                currency=currency,
                branch=request.user.branch,
                defaults={'balance': 0}  
            )
            logger.info(f"[Create Invoice]: {account_balance}")

            
            # accountts_receivable
            accounts_receivable, _ = ChartOfAccounts.objects.get_or_create(name="Accounts Receivable")

            # VAT rate
            vat_rate = VATRate.objects.get(status=True)

            # customer
            customer = Customer.objects.get(id=int(invoice_data['client_id'])) 
            
            # customer account
            customer_account = CustomerAccount.objects.get(customer=customer)

            # customer Account + Balances
            customer_account_balance, _ = CustomerAccountBalances.objects.get_or_create(
                account=customer_account,
                currency=currency, 
                defaults={'balance': 0}
            )
            
            amount_paid = Decimal(invoice_data['amount_paid'])
            invoice_total_amount = Decimal(invoice_data['payable'])
            
            # check for due invoices 
            # if Invoice.objects.filter(customer=customer, payment_status='Partial', branch=request.user.branch, currency=currency).exists():
            #     due_invoice = Invoice.objects.filter(customer=customer, payment_status='Partial', branch=request.user.branch).last()
            #     if amount_paid > due_invoice.amount_due:
            #         due_invoice.amount_paid += due_invoice.amount_due
            #         amount_paid -= due_invoice.amount_due
            #         due_invoice.payment_status= due_invoice.PaymentStatus.PAID
                    
            #         # refactor
            #         Payment.objects.create(
            #             invoice=due_invoice,
            #             amount_paid=due_invoice.amount_due,
            #             payment_method=invoice_data['payment_method'],
            #             amount_due= 0,
            #             user=request.user
            #         )
                    
            #         customer_account_balance.balance -= due_invoice.amount_due
            #         due_invoice.amount_due = 0
                    
            #     else:
            #         due_invoice.amount_due -= amount_paid
            #         due_invoice.amount_paid += amount_paid
                    
            #         Payment.objects.create(
            #             invoice=due_invoice,
            #             amount_paid=amount_paid,
            #             payment_method=invoice_data['payment_method'],
            #             user=request.user
            #         )
        
            #         customer_account_balance.balance -= amount_paid
            #         amount_paid = 0
                
            #     customer_account_balance.save()
            #     due_invoice.save()

            # calculate amount due
            amount_due = invoice_total_amount - amount_paid  
            
            with transaction.atomic():
                
                recurrence_period = int(invoice_data['next_due_days']) if invoice_data['next_due_days'] else 0
                
                invoice = Invoice.objects.create(
                    invoice_number=Invoice.generate_invoice_number(request.user.branch.name),  
                    customer=customer,
                    issue_date=timezone.now(),
                    due_date=timezone.now() + timezone.timedelta(days=int(invoice_data['days'])),
                    amount=invoice_total_amount,
                    amount_paid=amount_paid,
                    amount_due=amount_due,
                    discount_amount=invoice_data['discount'],
                    delivery_charge=invoice_data['delivery'],
                    vat=Decimal(invoice_data['vat_amount']),
                    payment_status = Invoice.PaymentStatus.PARTIAL if amount_due > 0 else Invoice.PaymentStatus.PAID,
                    branch = request.user.branch,
                    user=request.user,
                    currency=currency,
                    subtotal=invoice_data['subtotal'],
                    note=invoice_data['note'],
                    reocurring = invoice_data['recourring'],
                    products_purchased = ', '.join([f'{item['product_name']} x {item['quantity']} ' for item in items_data]),
                    recurrence_period = recurrence_period ,
                    next_due_date = datetime.datetime.now() + timedelta(days=recurrence_period )
                )
                
                # #create transaction
                transaction_obj = Transaction.objects.create(
                    date=timezone.now(),
                    description=invoice_data['note'],
                    account=accounts_receivable,
                    debit=Decimal(invoice_data['payable']),
                    credit=Decimal('0.00'),
                    customer=customer
                )

                # Cost of sales parent object
                cogs = COGS.objects.create(amount=Decimal(0))
                
                # Create InvoiceItem objects
                for item_data in items_data:
                    item = Inventory.objects.get(pk=item_data['inventory_id'])
                    product = Product.objects.get(pk=item.product.id)
                    
                    item.quantity -= item_data['quantity']
                    item.save()
                  
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        item=item,
                        quantity=item_data['quantity'],
                        unit_price=item_data['price'],
                        vat_rate = vat_rate
                    )
                    
                    # Create StockTransaction for each sold item
                    stock_transaction = StockTransaction.objects.create(
                        item=product,
                        transaction_type=StockTransaction.TransactionType.SALE,
                        quantity=item_data['quantity'],
                        unit_price=item.price,
                        invoice=invoice,
                        date=timezone.now()
                    )

                    # cost of sales item
                    COGSItems.objects.create(
                        invoice=invoice,
                        cogs=cogs,
                        product= Inventory.objects.get(product=product)

                    )
                    
                    # stock log  
                    ActivityLog.objects.create(
                        branch=request.user.branch,
                        inventory=item,
                        user=request.user,
                        quantity=item_data['quantity'],
                        total_quantity = item.quantity,
                        action='Sale',
                        invoice=invoice
                    )
                    
                # # Create VATTransaction
                VATTransaction.objects.create(
                    invoice=invoice,
                    vat_type=VATTransaction.VATType.OUTPUT,
                    vat_rate=VATRate.objects.get(status=True).rate,
                    tax_amount=invoice_data['vat_amount']
                )                                                          
                # Create Sale object
                sale = Sale.objects.create(
                    date=timezone.now(),
                    transaction=invoice,
                    total_amount=invoice_total_amount
                )
                
                #payment
                Payment.objects.create(
                    invoice=invoice,
                    amount_paid=amount_paid,
                    payment_method=invoice_data['payment_method'],
                    amount_due=invoice_total_amount - amount_paid,
                    user=request.user
                )


                # calculate total cogs amount
                cogs.amount = COGSItems.objects.filter(cogs=cogs, cogs__date=datetime.datetime.today())\
                                               .aggregate(total=Sum('product__cost'))['total'] or 0
                logger.info(f'COGS amount: {cogs.amount}')
                cogs.save()
                
                # updae account balance
                if invoice.payment_status == 'Partial':
                    customer_account_balance.balance += -amount_due
                    customer_account_balance.save()
                    
                # Update customer balance
                account_balance.balance = Decimal(invoice_data['payable']) + Decimal(account_balance.balance)
                account_balance.save()
                
                # return redirect('finance:invoice_preview', invoice.id)
                return JsonResponse({'success':True, 'invoice_id': invoice.id})

        except (KeyError, json.JSONDecodeError, Customer.DoesNotExist, Inventory.DoesNotExist) as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return render(request, 'finance/invoices/add_invoice.html')

@login_required
@transaction.atomic
def invoice_returns(request, invoice_id): # dont forget the payments
    invoice = get_object_or_404(Invoice, id=invoice_id)
    account = get_object_or_404(CustomerAccount, customer=invoice.customer)
    customer_account_balance = get_object_or_404(CustomerAccountBalances, account=account, currency=invoice.currency)

    sale = get_object_or_404(Sale, transaction=invoice)
    invoice_payment = get_object_or_404(Payment, invoice=invoice)
    stock_transactions = invoice.stocktransaction_set.all()  
    vat_transaction = get_object_or_404(VATTransaction, invoice=invoice)

    if invoice.payment_status == Invoice.PaymentStatus.PARTIAL:
        customer_account_balance.balance -= invoice.amount_due

    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    account = get_object_or_404(
        Account, 
        name=f"{request.user.branch} {invoice.currency.name} {invoice_payment.payment_method.capitalize()} Account", 
        type=account_types.get(invoice_payment.payment_method, None) 
    )
    account_balance = get_object_or_404(AccountBalance, account=account, currency=invoice.currency, branch=request.user.branch)
    account_balance.balance -= invoice.amount_paid

    for stock_transaction in stock_transactions:
        product = Inventory.objects.get(product=stock_transaction.item, branch=request.user.branch)
        product.quantity += stock_transaction.quantity
        product.save()

        ActivityLog.objects.create(
            invoice=invoice,
            product_transfer=None,
            branch=request.user.branch,
            user=request.user,
            action='returns',
            inventory=product,
            quantity=stock_transaction.quantity,
            total_quantity=product.quantity
        )

    InvoiceItem.objects.filter(invoice=invoice).delete() 
    StockTransaction.objects.filter(invoice=invoice).delete()
    
    account_balance.save()
    customer_account_balance.save()
    sale.delete()
    vat_transaction.delete()
    invoice.invoice_return=True
    invoice.save()

    return JsonResponse({'message': f'Invoice {invoice.invoice_number} successfully deleted'})
    

@login_required
@transaction.atomic
def delete_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    account = get_object_or_404(CustomerAccount, customer=invoice.customer)
    customer_account_balance = get_object_or_404(CustomerAccountBalances, account=account, currency=invoice.currency)

    sale = get_object_or_404(Sale, transaction=invoice)
    invoice_payment = get_object_or_404(Payment, invoice=invoice)
    stock_transactions = invoice.stocktransaction_set.all()  
    vat_transaction = get_object_or_404(VATTransaction, invoice=invoice)

    if invoice.payment_status == Invoice.PaymentStatus.PARTIAL:
        customer_account_balance.balance -= invoice.amount_due

    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    account = get_object_or_404(
        Account, 
        name=f"{request.user.branch} {invoice.currency.name} {invoice_payment.payment_method.capitalize()} Account", 
        type=account_types.get(invoice_payment.payment_method, None)  
    )
    account_balance = get_object_or_404(AccountBalance, account=account, currency=invoice.currency, branch=request.user.branch)
    account_balance.balance -= invoice.amount_paid

    for stock_transaction in stock_transactions:
        product = Inventory.objects.get(product=stock_transaction.item, branch=request.user.branch)
        product.quantity += stock_transaction.quantity
        product.save()

        ActivityLog.objects.create(
            invoice=invoice,
            product_transfer=None,
            branch=request.user.branch,
            user=request.user,
            action='sale cancelled',
            inventory=product,
            quantity=stock_transaction.quantity,
            total_quantity=product.quantity
        )

    InvoiceItem.objects.filter(invoice=invoice).delete() 
    StockTransaction.objects.filter(invoice=invoice).delete()
    
    account_balance.save()
    customer_account_balance.save()
    sale.delete()
    vat_transaction.delete()
    invoice.cancelled=True
    invoice.save()

    return JsonResponse({'message': f'Invoice {invoice.invoice_number} successfully deleted'})
    
    
    
@login_required       
def invoice_details(request, invoice_id):
    invoice = Invoice.objects.filter(id=invoice_id, branch=request.user.branch).values(
        'invoice_number',
        'customer__id', 
        'customer__name', 
        'products_purchased', 
        'payment_status', 
        'amount'
    )
    return JsonResponse(list(invoice), safe=False)


@login_required
def customer(request):
    if request.method == 'GET':
        customers = Customer.objects.all().values()
        return JsonResponse(list(customers), safe=False)
    
    elif request.method == 'POST':
        data = json.loads(request.body)
        
        validation_errors = validate_customer_data(data)
        if validation_errors:
            return JsonResponse({'success': False, 'message': 'Validation errors occurred:', 'errors': validation_errors})
    
        if Customer.objects.filter(phone_number=data['phonenumber']).exists():
            return JsonResponse({'success': False, 'message': 'Customer exists'})
        else:
            customer = Customer.objects.create(
                name=data['name'],
                email=data['email'],
                address=data['address'],
                phone_number=data['phonenumber'],
                branch=request.user.branch
            )
            account = CustomerAccount.objects.create(customer=customer)

        balances_to_create = [
            CustomerAccountBalances(account=account, currency=currency, balance=0) 
            for currency in Currency.objects.all()
        ]
        CustomerAccountBalances.objects.bulk_create(balances_to_create)

        return JsonResponse({'success': True, 'message': 'Customer successfully created'})

    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})

def validate_customer_data(data):
    errors = {}
    if 'name' not in data or len(data['name']) < 2:
        errors['name'] = 'Name is required and must be at least 2 characters long.'

    if 'email' not in data or not validate_email(data['email']):
        errors['email'] = 'A valid email address is required.'

    if 'address' not in data:
        errors['address'] = 'Address is required.'

    if 'phonenumber' not in data:
        errors['phonenumber'] = 'Phone number is required.'

    return errors

def validate_email(email):
    import re
    email_regex = r"^[a-zA-Z0-9.!#$%&'*+/=?^_`{|}~-]+@[a-zA-Z0-9-]+(?:\.[a-zA-Z0-9-]+)*$"
    return bool(re.match(email_regex, email))

@login_required
def customer_list(request):
    search_query = request.GET.get('q', '')
    
    customers = Customer.objects.filter(branch=request.user.branch)
    accounts = CustomerAccountBalances.objects.all()
    
    total_balances_per_currency = CustomerAccountBalances.objects.filter(account__customer__branch=request.user.branch).values('currency__name').annotate(
        total_balance=Sum('balance')
    )
    
    if search_query:
        customers = CustomerAccount.objects.filter(Q(customer__name__icontains=search_query))
        
    if 'receivable' in request.GET:
        negative_balances_per_currency = CustomerAccountBalances.objects.filter(account__customer__branch=request.user.branch, balance__lt=0) \
            .values('currency') \
            .annotate(total_balance=Sum('balance'))

        customers = Customer.objects.filter(
            id__in=negative_balances_per_currency.values('account__customer_id'),
        ).distinct()
        
        total_balances_per_currency = negative_balances_per_currency.values('currency__name').annotate(
            total_balance=Sum('balance')
        )
        
        logger.info(f'Customers:{total_balances_per_currency.values}')

    if 'download' in request.GET: 
        customers = Customer.objects.all() 
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=customers.xlsx'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for col_num, header_title in enumerate(['Customer Name', 'Phone Number', 'Email', 'Account Balance'], start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.alignment = header_alignment
            
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(len(header_title), 20)

        customer_accounts = CustomerAccountBalances.objects.all()
        for customer in customer_accounts:
            worksheet.append(
                [
                    customer.account.customer.name, 
                    customer.account.customer.phone_number, 
                    customer.account.customer.email, 
                    customer.balance if customer.balance else 0,
                ]
            )  
            
        workbook.save(response)
        return response
        
    return render(request, 'finance/customers/customers.html', {
        'customers':customers, 
        'accounts':accounts,
        'total_balances_per_currency':total_balances_per_currency,
    })

@login_required
def update_customer(request, customer_id):
    customer = get_object_or_404(Customer, pk=customer_id)

    if request.method == 'POST':  
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f'{customer.name} details updated successfully')  #
            return redirect('finance:customer_list')  
    else:
        form = CustomerForm(instance=customer)  

    return render(request, 'finance/customers/update_customer.html', {'form': form, 'customer': customer}) 

def delete_customer(request, customer_id):
    if request.method == 'DELETE':
        customer = get_object_or_404(Customer, pk=customer_id)

        customer_name = customer.name  
        customer.delete()
        messages.success(request, f'{customer_name} deleted successfully.')
        return JsonResponse({'status': 'success', 'message': f'Customer {customer_name} deleted successfully.'})  
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})  
    

@login_required
def customer_account(request, customer_id):
    form = customerDepositsForm()
    refund_form = customerDepositsRefundForm()
    customer = get_object_or_404(Customer, id=customer_id)

    account = CustomerAccountBalances.objects.filter(account__customer=customer)

    invoices = Invoice.objects.filter(
        customer=customer, 
        branch=request.user.branch, 
        status=True
    )
    
    invoice_payments = Payment.objects.filter(
        invoice__branch=request.user.branch, 
        invoice__customer=customer
    ).order_by('-payment_date')

    filters = Q()
    if request.GET.get('q'):
        filters &= Q(payment_status=request.GET['q'])
    if request.GET.get('search_query'):
        search_query = request.GET['search_query']
        filters &= (Q(invoice_number__icontains=search_query) | Q(issue_date__icontains=search_query))

    invoices = invoices.filter(filters)

    if request.GET.get('email_bool'):
        send_account_statement_email(customer.id, request.user.branch.id, request.user.id)
        return JsonResponse({'message': 'Email sent'})

    return render(request, 'finance/customer.html', {
        'form':form,
        'account': account,
        'invoices': invoices,
        'customer': customer,
        'refund_form':refund_form,
        'invoice_count': invoices.count(),
        'invoice_payments': invoice_payments,
        'paid': invoices.filter(payment_status='Paid').count(),  
        'due': invoices.filter(payment_status='Partial').count(), 
    })


@login_required
@transaction.atomic
def add_customer_deposit(request, customer_id):
    # payload
    """
        customer_id
        amount
        currency
        payment_method
        reason
        payment_reference
    """
    
    try: 
        # get payload
        data = json.loads(request.body)
        customer_id = data.get('customer_id')
        amount = data.get('amount')
        currency = data.get('currency')
        payment_method = data.get('payment_method')
        reason = data.get('reason')
        payment_reference = data.get('payment_reference')        
        
        # payment_reference validation
        if CustomerDeposits.objects.filter(payment_reference=payment_reference).exists():
            return JsonResponse(
                {
                    'success':False,
                    'message': f'Payment reference: {payment_reference} exists'
                }
            )   
                                                   
        # get currency
        currency = Currency.objects.get(id=currency)
        
        # get account types
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }
        
        account_name = f"{request.user.branch} {currency.name} {payment_method.capitalize()} Account"
        
        
        account, _ = Account.objects.get_or_create(name=account_name, type=account_types[payment_method])
        
        # get or create the account balances
        account_balance, _ = AccountBalance.objects.get_or_create(
            account=account,
            currency=currency,
            branch=request.user.branch,
            defaults={'balance': 0}  
        )
        
        account_balance.balance += Decimal(amount)
        account_balance.save()
        logger.info(f"[FINANCE]: deposit -> System {account}")
        
        # check if customer exits
        customer = get_object_or_404(Customer, id=customer_id)  
        logger.info(f"[FINANCE]: deposit -> customer {customer}")
        customer_account = CustomerAccount.objects.get(customer=customer)
        
        customer_account_bal_object, _ = CustomerAccountBalances.objects.get_or_create(
                account=customer_account,
                currency=currency, 
                defaults={'balance': 0}
        )  
        logger.info(f"[FINANCE]: deposit -> customer account object {customer_account_bal_object}")
        
        # effect customer deposit
        customer_deposit = CustomerDeposits.objects.create(
            customer_account=customer_account_bal_object,
            amount=amount,
            currency=currency,
            payment_method=payment_method,
            reason=reason,
            payment_reference=payment_reference,
            cashier=request.user,
            branch=request.user.branch
        )
        
        # effect customer account balances
        customer_account_bal_object.balance += amount
        
        customer_account_bal_object.save()
        
        Cashbook.objects.create(
            issue_date=customer_deposit.date_created,
            description=f'{customer_deposit.payment_method.upper()} deposit ({customer_deposit.customer_account.account.customer.name})',
            debit=True,
            credit=False,
            amount=customer_deposit.amount,
            currency=customer_deposit.currency,
            branch=customer_deposit.branch
        )

        return JsonResponse(
            {
                "success":True,
                "message": f"Customer Deposit of {currency} {amount:2f} has been successfull",
            },
            status=200
        )
    except Exception as e:
        return JsonResponse(
            {
                "message": f"{e}",
                'success':False
            },status=500)


@login_required    
def deposits_list(request):
    deposits = CustomerDeposits.objects.filter(branch=request.user.branch).order_by('-date_created')
    return render(request, 'finance/deposits.html', {
        'deposits':deposits,
        'total_deposits': deposits.aggregate(Sum('amount'))['amount__sum'] or 0,
    })

@login_required
@transaction.atomic
def refund_customer_deposit(request, deposit_id):
    try:
        deposit = CustomerDeposits.objects.get(id=deposit_id)
    except CustomerDeposits.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Deposit not found'}, status=404)
    
    try:
        data = json.loads(request.body)
        amount = Decimal(data.get('amount', 0))
        if amount <= 0:
            return JsonResponse({'success': False, 'message': 'Invalid amount'}, status=400)
    except (json.JSONDecodeError, TypeError, ValueError):
        return JsonResponse({'success': False, 'message': 'Invalid input data'}, status=400)

    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    account_name = f"{request.user.branch} {deposit.currency.name} {deposit.payment_method.capitalize()} Account"

    try:
        account = Account.objects.get(name=account_name, type=account_types[deposit.payment_method])
        account_balance = AccountBalance.objects.get(
            account=account,
            currency=deposit.currency,
            branch=request.user.branch,
        )
    except (Account.DoesNotExist, AccountBalance.DoesNotExist) as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    if amount > deposit.amount:
        return JsonResponse({'success': False, 'message': 'Refund amount exceeds deposit amount'}, status=400)
    
    account_balance.balance -= amount
    diff_amount = deposit.amount - amount

    if diff_amount == 0:
        deposit.delete()
    else:
        deposit.amount = diff_amount
        deposit.save()

    Cashbook.objects.create(
        issue_date=datetime.date.today(),
        description=f'{deposit.payment_method.upper()} deposit refund ({deposit.customer_account.account.customer.name})',
        debit=False,
        credit=True,
        amount=amount,
        currency=deposit.currency,
        branch=deposit.branch
    )

    account_balance.save()

    return JsonResponse({'success': True}, status=200)

        
@login_required
@transaction.atomic
def edit_customer_deposit(request, deposit_id):
    try:
        deposit = CustomerDeposits.objects.get(id=deposit_id)
    except CustomerDeposits.DoesNotExist:
        messages.warning(request, 'Deposit not found')
        return redirect('finance:customer_account', deposit.customer_account.account.customer.id)
    
    if request.method == 'POST':
        form = customerDepositsForm(request.POST)
        if not form.is_valid():
            messages.warning(request, 'Invalid form submission')
            return redirect('finance:edit_customer_deposit', deposit_id)

        amount = Decimal(request.POST.get('amount'))
        if amount <= 0:
            messages.warning(request, 'Amount cannot be zero or negative')
            return redirect('finance:edit_customer_deposit', deposit_id)

        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }

        account_name = f"{request.user.branch} {deposit.currency.name} {deposit.payment_method.capitalize()} Account"
        
        try:
            account = Account.objects.get(name=account_name, type=account_types[deposit.payment_method])
            account_balance = AccountBalance.objects.get(
                account=account,
                currency=deposit.currency,
                branch=request.user.branch,
            )
        except (Account.DoesNotExist, AccountBalance.DoesNotExist) as e:
            messages.warning(request, str(e))
            return redirect('finance:edit_customer_deposit', deposit_id)
        
        adj_amount = amount - deposit.amount

        if adj_amount != 0:
            if adj_amount > 0:
                account_balance.balance += adj_amount
                debit, credit = True, False
            else:
                account_balance.balance += adj_amount 
                debit, credit = False, True

            Cashbook.objects.create(
                issue_date=datetime.date.today(),
                description=f'{deposit.payment_method.upper()} deposit adjustment ({deposit.customer_account.account.customer.name})',
                debit=debit,
                credit=credit,
                amount=abs(adj_amount),
                currency=deposit.currency,
                branch=deposit.branch
            )

            account_balance.save()
            deposit.amount = amount
            deposit.save()
            messages.success(request, 'Customer deposit successfully updated')
            return redirect('finance:customer', deposit.customer_account.account.customer.id)
    else:
        form = customerDepositsForm(instance=deposit)

    return render(request, 'finance/customers/edit_deposit.html', {'form': form})
    

@login_required
def customer_deposits(request): 
    customer_id = request.GET.get('customer_id')
    
    if customer_id: 
        deposits = CustomerDeposits.objects.filter(branch=request.user.branch).values(
            'customer_account__account__customer_id',
            'date_created',
            'amount', 
            'reason',
            'currency__name', 
            'currency__symbol', 
            'payment_method',
            'payment_reference',
            'cashier__username', 
            'id'
        ).order_by('-date_created')
        return JsonResponse(list(deposits), safe=False)
    else:
        return JsonResponse({
            'success':False,
            'message':f'{customer_id} was not provided'
        })

@login_required
def customer_account_transactions_json(request):
    customer_id = request.GET.get('customer_id')
    transaction_type = request.GET.get('type')

    customer = get_object_or_404(Customer, id=customer_id)  

    if transaction_type == 'invoices':
        invoices = Invoice.objects.filter(
            customer=customer, 
            branch=request.user.branch, 
            status=True
        ).order_by('-issue_date').values(
            'issue_date',
            'invoice_number',
            'products_purchased', 
            'amount_paid', 
            'amount_due', 
            'amount', 
            'user__username',
            'payment_status'
        )
        return JsonResponse(list(invoices), safe=False)
    else:
        return JsonResponse({'message': 'Invalid transaction type.'}, status=400)  

@login_required
def customer_account_payments_json(request):
    customer_id = request.GET.get('customer_id')
    transaction_type = request.GET.get('type')

    customer = get_object_or_404(Customer, id=customer_id)

    if transaction_type == 'invoice_payments':
        invoice_payments = Payment.objects.select_related('invoice', 'invoice__currency', 'user').filter(
            invoice__branch=request.user.branch, 
            invoice__customer=customer
        ).order_by('-payment_date').values(
            'invoice__products_purchased',
            'payment_date',
            'invoice__invoice_number',
            'invoice__currency__symbol', 
            'invoice__payment_status',
            'invoice__amount_due',
            'invoice__amount', 
            'user__username', 
            'amount_paid', 
            'amount_due'
        )
        return JsonResponse(list(invoice_payments), safe=False)
    else:
        return JsonResponse({'message': 'Invalid transaction type.'}, status=400)  


@login_required
def customer_account_json(request, customer_id):
    account = CustomerAccountBalances.objects.filter(account__customer__id=customer_id).values(
        'currency__symbol', 'balance'
    )   
    return JsonResponse(list(account), safe=False)

@login_required
def print_account_statement(request, customer_id):
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        
        account = CustomerAccountBalances.objects.filter(account__customer=customer)
        
        invoices = Invoice.objects.filter(
            customer=customer, 
            branch=request.user.branch, 
            status=True
        )
    except:
        messages.warning(request, 'Error in processing the request')
        return redirect('finance:customer')

    invoice_payments = Payment.objects.select_related('invoice', 'invoice__currency', 'user').filter(
        invoice__branch=request.user.branch, 
        invoice__customer=customer
    ).order_by('-payment_date')
    
    return render(request, 'finance/customers/print_customer_statement.html', {
        'customer':customer,
        'account':account,
        'invoices':invoices, 
        'invoice_payments':invoice_payments
    })

# currency views  
@login_required  
def currency(request):
    return render(request, 'finance/currency/currency.html')

@login_required
def currency_json(request):
    currency_id = request.GET.get('id', '')
    currency = Currency.objects.filter(id=currency_id).values()
    return JsonResponse(list(currency), safe=False)


@login_required
def add_currency(request):
    if request.method == 'POST':
        form = CurrencyForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Currency added successfully!')  
            except Exception as e: 
                messages.error(request, f'Error adding currency: {e}')
            return redirect('finance:currency') 
    else:
        form = CurrencyForm()

    return render(request, 'finance/currency/currency_add.html', {'form': form})


@login_required
def update_currency(request, currency_id):
    currency = get_object_or_404(Currency, id=currency_id)  

    if request.method == 'POST': 
        form = CurrencyForm(request.POST, instance=currency)  
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Currency updated successfully') 
            except Exception as e: 
                messages.error(request, f'Error updating currency: {e}')
            return redirect('finance:currency')  
    else:
        form = CurrencyForm(instance=currency) 

    return render(request, 'finance/currency/currency_add.html', {'form': form})

@login_required
def delete_currency(request, currency_id):
    if request.method == 'POST': 
        currency = get_object_or_404(Currency, id=currency_id)
        
        try:
            if currency.invoice_set.exists() or currency.accountbalance_set.exists() or currency.expense_set.exists():  
                raise Exception("Currency is in use and cannot be deleted.")

            currency.delete()
            return JsonResponse({'message': 'Currency deleted successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'message':'Deletion Failed'})


@login_required
def finance_settings(request):
    return render(request, 'finance/settings/settings.html')
    
# Reports
@login_required
def expenses_report(request):
    
    template_name = 'finance/reports/expenses.html'
    
    search = request.GET.get('search', '')
    start_date_str = request.GET.get('startDate', '')
    end_date_str = request.GET.get('endDate', '')
    category_id = request.GET.get('category', '')
   
    if start_date_str and end_date_str:
        try:
            end_date = datetime.date.fromisoformat(end_date_str)
            start_date = datetime.date.fromisoformat(start_date_str)
        except ValueError:
            return JsonResponse({'messgae':'Invalid date format. Please use YYYY-MM-DD.'})
    else:
        start_date = ''
        end_date= ''
        
    try:
        category_id = int(category_id) if category_id else None
    except ValueError:
        return JsonResponse({'messgae':'Invalid category or search ID.'})

    expenses = Expense.objects.all()  
    
    if search:
        expenses = expenses.filter(Q('amount=search'))
    if start_date:
        start_date = parse_date(start_date_str)
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        end_date = parse_date(end_date_str)
        expenses = expenses.filter(date__lte=end_date)
    if category_id:
        expenses = expenses.filter(category__id=category_id)
    
    return generate_pdf(
        template_name,
        {
            'title': 'Expenses', 
            'date_range': f"{start_date} to {end_date}", 
            'report_date': datetime.date.today(),
            'total_expenses':calculate_expenses_totals(expenses),
            'expenses':expenses
        }
    )


@login_required 
def invoice_preview(request, invoice_id):
    invoice = Invoice.objects.get(id=invoice_id)
    invoice_items = InvoiceItem.objects.filter(invoice=invoice)
    return render(request, 'Pos/printable_receipt.html', {'invoice_id':invoice_id, 'invoice':invoice, 'invoice_items':invoice_items})


@login_required
def invoice_preview_json(request, invoice_id):
    try:
        invoice = Invoice.objects.get(id=invoice_id)

    except Invoice.DoesNotExist:
        return JsonResponse({"error": "Invoice not found"}, status=404) 
     
    invoice_items = InvoiceItem.objects.filter(invoice=invoice).values(
        'item__product__name', 
        'quantity',
        'item__product__description',
        'total_amount'
    )

    invoice_dict = {
        field.name: getattr(invoice, field.name)
        for field in invoice._meta.fields
        if field.name not in ['customer', 'currency', 'branch', 'user']
    }

    invoice_dict['customer_name'] = invoice.customer.name
    invoice_dict['customer_email'] = invoice.customer.email 
    invoice_dict['currency_symbol'] = invoice.currency.symbol
    invoice_dict['amount_paid'] = invoice.amount_paid
    
    if invoice.branch:
        invoice_dict['branch_name'] = invoice.branch.name
        invoice_dict['branch_phone'] = invoice.branch.phonenumber
        invoice_dict['branch_email'] = invoice.branch.email
        
    invoice_dict['user_username'] = invoice.user.username  
    
    invoice_data = {
        'invoice': invoice_dict,
        'invoice_items': list(invoice_items),
    }

    return JsonResponse(invoice_data)

@login_required
def invoice_pdf(request):
    template_name = 'finance/reports/invoice.html'
    invoice_id = request.GET.get('id', '')
    if invoice_id:
        try:
            invoice = get_object_or_404(Invoice, pk=invoice_id)

            invoice_items = InvoiceItem.objects.filter(invoice=invoice)
            
        except Invoice.DoesNotExist:
            return HttpResponse("Invoice not found")
    else:
        return HttpResponse("Invoice ID is required")
    
    return generate_pdf(
        template_name,
        {
            'title': 'Invoice', 
            'report_date': datetime.date.today(),
            'invoice':invoice,
            'invoice_items':invoice_items
        }
    )
    
# emails
@login_required
def send_invoice_email(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        invoice_id = data['invoice_id']
        invoice = Invoice.objects.get(id=invoice_id)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        account = CustomerAccount.objects.get(customer__id = invoice.customer.id)
        
        html_string = render_to_string('Pos/receipt.html', {'invoice': invoice, 'invoice_items':invoice_items, 'account':account})
        buffer = BytesIO()

        pisa.CreatePDF(html_string, dest=buffer) 

        email = EmailMessage(
            'Your Invoice',
            'Please find your invoice attached.',
            'your_email@example.com',
            ['recipient_email@example.com'],
        )
        
        buffer.seek(0)
        email.attach(f'invoice_{invoice.invoice_number}.pdf', buffer.getvalue(), 'application/pdf')

        # Send the email
        email.send()

        task = send_invoice_email_task.delay(data['invoice_id']) 
        task_id = task.id 
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=invoice_{invoice.invoice_number}.pdf'
        
        return response
    return JsonResponse({'success': False, 'error':'error'})


#whatsapp
@login_required
def send_invoice_whatsapp(request, invoice_id):
    try:
        
        invoice = Invoice.objects.get(pk=invoice_id)
        invoice_items = InvoiceItem.objects.filter(invoice=invoice)
        img = settings.STATIC_URL + "/assets/logo.png"
    
        html_string = render_to_string('Pos/invoice_template.html', {'invoice': invoice, 'request':request, 'invoice_items':invoice_items, 'img':img})
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        if not pisa_status.err:
          
            s3 = boto3.client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            invoice_filename = f"invoice_{invoice.invoice_number}.pdf"
            s3.put_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                Key=f"invoices/{invoice_filename}",
                Body=pdf_buffer.getvalue(),
                ContentType="application/pdf",
                ACL="public-read",
            )
            s3_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/invoices/{invoice_filename}"

            account_sid = 'AC6890aa7c095ce1315c4a3a86f13bb403'
            auth_token = '897e02139a624574c5bd175aa7aaf628'
            client = Client(account_sid, auth_token)
            from_whatsapp_number = 'whatsapp:' + '+14155238886'
            to_whatsapp_number = 'whatsapp:' + '+263778587612'

            message = client.messages.create(
                from_=from_whatsapp_number,
                body="Your invoice is attached.",
                to=to_whatsapp_number,
                media_url=s3_url
            )
            logger.info(f"WhatsApp message SID: {message.sid}")
            return JsonResponse({"success": True, "message_sid": message.sid})
        else:
            logger.error(f"PDF generation error for Invoice ID: {invoice_id}")
            return JsonResponse({"error": "PDF generation failed"})
    except Invoice.DoesNotExist:
        logger.error(f"Invoice not found with ID: {invoice_id}")
        return JsonResponse({"error": "Invoice not found"})
    except Exception as e:
        logger.exception(f"Error sending invoice via WhatsApp: {e}")
        return JsonResponse({"error": "Error sending invoice via WhatsApp"})
    

@login_required
def end_of_day(request):
    today = timezone.now().date()
    
    user_timezone_str = request.user.timezone if hasattr(request.user, 'timezone') else 'UTC'
    user_timezone = pytz_timezone(user_timezone_str)  

    # make a utility
    def filter_by_date_range(start_date, end_date):
        start_datetime = user_timezone.localize(
            timezone.datetime.combine(start_date, timezone.datetime.min.time())
        )
        end_datetime = user_timezone.localize(
            timezone.datetime.combine(end_date, timezone.datetime.max.time())
        )
        return Invoice.objects.filter(branch=request.user.branch, issue_date__range=[start_datetime, end_datetime])

    now = timezone.now().astimezone(user_timezone)
    today = now.date()

    now = timezone.now() 
    today = now.date()  
    
    invoices = filter_by_date_range(today, today)
    withdrawals = CashWithdraw.objects.filter(user__branch=request.user.branch, date=today, status=False)
    
    total_cash_amounts = [
        {
            'total_invoices_amount' : invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
            'total_withdrawals_amount' : withdrawals.aggregate(Sum('amount'))['amount__sum'] or 0
        }
    ]

    sold_inventory = (
        StockTransaction.objects
        .filter(invoice__branch=request.user.branch, date=today, transaction_type=StockTransaction.TransactionType.SALE)
        .values('item__id', 'item__name')
        .annotate(quantity_sold=Sum('quantity'))
    )
    
    if request.method == 'GET':
        all_inventory = Inventory.objects.filter(branch=request.user.branch, status=True).values(
            'id', 'product__name', 'quantity'
        )

        inventory_data = []
        for item in sold_inventory:
            sold_info = next((inv for inv in all_inventory if item['item__id'] == inv['id']), None)
            
            if sold_info:
                inventory_data.append({
                    'id': item['item__id'],
                    'name': item['item__name'],
                    'initial_quantity': item['quantity_sold'] + sold_info['quantity'] if sold_info else 0,
                    'quantity_sold':  item['quantity_sold'],
                    'remaining_quantity':sold_info['quantity'] if sold_info else 0,
                    'physical_count': None
                })
           
        return JsonResponse({'inventory': inventory_data, 'total_cash_amounts':total_cash_amounts})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            inventory_data = []
            
            for item in data:
                try:
                    inventory = Inventory.objects.get(id=item['item_id'], branch=request.user.branch, status=True)
                    inventory.physical_count = item['physical_count']
                    inventory.save()

                    sold_info = next((i for i in sold_inventory if i['item__id'] == inventory.id), None)
                    inventory_data.append({
                        'id': inventory.id,
                        'name': inventory.product.name,
                        'initial_quantity': inventory.quantity,
                        'quantity_sold': sold_info['quantity_sold'] if sold_info else 0,
                        'remaining_quantity': inventory.quantity - (sold_info['quantity_sold'] if sold_info else 0),
                        'physical_count': inventory.physical_count,
                        'difference': inventory.physical_count - (inventory.quantity - (sold_info['quantity_sold'] if sold_info else 0))
                    })
                except Inventory.DoesNotExist:
                    return JsonResponse({'success': False, 'error': f'Inventory item with id {item["item_id"]} does not exist.'})

            today_min = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
            today_max = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
            
            # Invoice data
            invoices = Invoice.objects.filter(branch=request.user.branch, issue_date__range=(today_min, today_max))
            partial_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PARTIAL)
            paid_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PAID, branch=request.user.branch)
           
            # Expenses
            expenses = Expense.objects.filter(branch=request.user.branch, date=today)
            confirmed_expenses = expenses.filter(status=True)
            unconfirmed_expenses = expenses.filter(status=False)
            
            # Accounts
            account_balances = AccountBalance.objects.filter(branch=request.user.branch)

            html_string = render_to_string('day_report.html', {
                'request':request,
                'invoices':invoices,
                'expenses':expenses,
                'date': today,
                'inventory_data': inventory_data,
                'total_sales': paid_invoices.aggregatea,
                'partial_payments': partial_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'total_paid_invoices': paid_invoices.count(),
                'total_partial_invoices': partial_invoices.count(),
                'total_expenses': confirmed_expenses.aggregate(Sum('amount'))['amount__sum'] or 0,
                'confirmed_expenses': confirmed_expenses,
                'unconfirmed_expenses': unconfirmed_expenses,
                'account_balances': account_balances,
            })
            
            pdf_buffer = BytesIO()
            pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
            if not pisa_status.err:
                filename = f"{request.user.branch.name}_today_report_{today}.pdf"
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "error": "Error generating PDF."})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data.'})
        except Exception as e:
            logger.exception(f"Error processing request: {e}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def invoice_payment_track(request):
    invoice_id = request.GET.get('invoice_id', '')
    
    if invoice_id:
        payments = Payment.objects.filter(invoice__id=invoice_id).order_by('-payment_date').values(
            'payment_date', 'amount_paid', 'payment_method', 'user__username'
        )
    return JsonResponse(list(payments), safe=False)

@login_required
def day_report(request, inventory_data):
    today_min = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_max = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # invoice data
    invoices = Invoice.objects.filter(branch=request.user.branch, issue_date__range=(today_min, today_max))
    
    partial_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PARTIAL)
    paid_invoices = invoices.filter(payment_status=Invoice.PaymentStatus.PAID)
    
    # expenses
    expenses = Expense.objects.filter(branch=request.user.branch, date=datetime.date.today())
    
    confirmed_expenses = expenses.filter(status=True)
    unconfirmed_expenses = expenses.filter(staus=False)
    
    # accounts
    account_balances = AccountBalance.objects.filter(branch=request.user.branch)
    
    try:
        html_string = render_to_string('day_report.html',{
                'request':request,
                'invoices':invoices,
                'date': datetime.date.today(),
                'inventory_data': inventory_data,
                'total_sales': paid_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'partial_payments': partial_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'total_paid_invoices': paid_invoices.count(),
                'expenses':expenses,
                'total_partial_invoices': partial_invoices.count(),
                'total_expenses': confirmed_expenses.aggregate(Sum('amount'))['amount__sum'] or 0,
                'confirmed_expenses': confirmed_expenses,
                'unconfirmed_expenses': unconfirmed_expenses,
                'account_balances': account_balances,
            })
        
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        if not pisa_status.err:
            # Save PDF to S3 and get URL
            s3 = boto3.client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            invoice_filename = f"{request.user.branch} today's ({datetime.date.today}) report.pdf"
            s3.put_object(
                Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                Key=f"daily_reports/{invoice_filename}",
                Body=pdf_buffer.getvalue(),
                ContentType="application/pdf",
                ACL="public-read",
            )
            s3_url = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.amazonaws.com/invoices/{invoice_filename}"

            # Send WhatsApp Message with Twilio
            account_sid = settings.TWILIO_ACCOUNT_SID
            auth_token = settings.TWILIO_AUTH_TOKEN
            client = Client(account_sid, auth_token)
            from_whatsapp_number = 'whatsapp:' + '+14155238886'
            to_whatsapp_number = 'whatsapp:' + '+263778587612'

            message = client.messages.create(
                from_=from_whatsapp_number,
                body="Today's report.",
                to=to_whatsapp_number,
                media_url=s3_url
            )

            logger.info(f"WhatsApp message SID: {message.sid}")
            return JsonResponse({"success": True, "message_sid": message.sid})
    except Exception as e:
        logger.exception(f"Error sending invoice via WhatsApp: {e}")
        return JsonResponse({"error": "Error sending invoice via WhatsApp"})

   
@login_required 
@transaction.atomic 
def cash_transfer(request):
    form = TransferForm()
    transfers = CashTransfers.objects.filter(branch=request.user.branch)
    
    account_types = {
        'cash': Account.AccountType.CASH,
        'bank': Account.AccountType.BANK,
        'ecocash': Account.AccountType.ECOCASH,
    }

    if request.method == 'POST':
        form = TransferForm(request.POST)
        
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.user = request.user
            transfer.notification_type = 'Expense'
            transfer.from_branch = request.user.branch
            transfer.branch = request.user.branch
            transfer.received_status = False
            
            account_name = f"{request.user.branch} {transfer.currency.name} {transfer.transfer_method.capitalize()} Account"

            try:
                account = Account.objects.get(name=account_name, type=account_types[transfer.transfer_method.lower()])
            except Account.DoesNotExist:
                messages.error(request, f"Account '{account_name}' not found.")
                return redirect('finance:cash_transfer')  

            try:
                account_balance = AccountBalance.objects.select_for_update().get(
                    account=account,
                    currency=transfer.currency,
                    branch=request.user.branch
                )
            except AccountBalance.DoesNotExist:
                messages.error(request, "Account balance record not found.")
                return redirect('finance:cash_transfer')

            if account_balance.balance < transfer.amount:
                messages.error(request, "Insufficient funds in the account.")
                return redirect('finance:cash_transfer')  

            account_balance.balance -= transfer.amount
            account_balance.save()
            transfer.save()  
            
            messages.success(request, 'Money successfully transferred.')
            return redirect('finance:cash_transfer')  
        else:
            messages.error(request, "Invalid form data. Please correct the errors.")
    return render(request, 'finance/transfers/cash_transfers.html', {'form': form, 'transfers':transfers})

@login_required
def finance_notifications_json(request):
    notifications = FinanceNotifications.objects.filter(status=True).values(
        'transfer__id', 
        'transfer__to',
        'expense__id',
        'expense__branch',
        'invoice__id',
        'invoice__branch',
        'notification',
        'notification_type',
        'id'
    )
    return JsonResponse(list(notifications), safe=False)


@login_required
@transaction.atomic
def cash_transfer_list(request):
    search_query = request.GET.get('q', '')
    transfers = CashTransfers.objects.filter(to=request.user.branch.id)
    
    if search_query:
        transfers = transfers.filter(Q(date__icontains=search_query))
        
    return render(request, 'finance/transfers/cash_transfers_list.html', {'transfers':transfers, 'search_query':search_query})

@login_required
@transaction.atomic
def receive_money_transfer(request, transfer_id):
    if transfer_id:
        transfer = get_object_or_404(CashTransfers, id=transfer_id)
        account_types = {
            'cash': Account.AccountType.CASH,
            'bank': Account.AccountType.BANK,
            'ecocash': Account.AccountType.ECOCASH,
        }
        
        account_name = f"{request.user.branch} {transfer.currency.name} {transfer.transfer_method.capitalize()} Account"

        try:
            account, _ = Account.objects.get_or_create(name=account_name, type=account_types[transfer.transfer_method.lower()])
        except Account.DoesNotExist:
            return JsonResponse({'message':f"Account '{account_name}' not found."}) 

        try:
            account_balance, _ = AccountBalance.objects.get_or_create(
                account=account,
                currency=transfer.currency,
                branch=request.user.branch
            )
        except AccountBalance.DoesNotExist:
            messages.error(request, )
            return JsonResponse({'message':"Account balance record not found."})  

        account_balance.balance += transfer.amount
        account_balance.save()
        
        transfer.received_status = True
        transfer.save() 
        return JsonResponse({'message':True})  
    return JsonResponse({'message':"Transfer ID is needed"})  


@login_required
@transaction.atomic
def create_quotation(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        qoute_data = data['data'][0]  
        items_data = data['items']
        
        customer = Customer.objects.get(id=int(qoute_data['client_id']))
        currency = Currency.objects.get(id=qoute_data['currency'])
        
        qoute = Qoutation.objects.create(
            customer = customer,
            amount =  Decimal(qoute_data['subtotal']),
            branch = request.user.branch,
            currency = currency,
            qoute_reference = Qoutation.generate_qoute_number(request.user.branch.name),
            products = ', '.join([f'{item['product_name']} x {item['quantity']}' for item in items_data])
        )
        
        for item_data in items_data:
            item = Inventory.objects.get(pk=item_data['inventory_id'])
            
            QoutationItems.objects.create(
                qoute=qoute,
                product=item,
                unit_price=item.price,
                quantity=item_data['quantity'],
                total_amount= item.price * item_data['quantity'],
            )
        return JsonResponse({'success': True, 'qoute_id': qoute.id})
    return JsonResponse({'success': False})

@login_required        
def qoutation_list(request):
    search_query = request.GET.get('q', '')
    qoutations = Qoutation.objects.filter(branch=request.user.branch).order_by('-date')
 
    if search_query:

        qoutations = qoutations.filter(
            Q(customer__name__icontains=search_query)|
            Q(products__icontains=search_query)|
            Q(date__icontains=search_query)|
            Q(qoute_reference__icontains=search_query)
        )
        
    return render(request, 'finance/qoutations.html', {'qoutations':qoutations, 'search_query':search_query})
        
@login_required 
def qoute_preview(request, qoutation_id):
    qoute = Qoutation.objects.get(id=qoutation_id)
    qoute_items = QoutationItems.objects.filter(qoute=qoute)
    return render(request, 'Pos/qoute.html', {'qoute':qoute, 'qoute_items':qoute_items})

@login_required
def delete_qoute(request, qoutation_id):
    qoute = get_object_or_404(Qoutation, id=qoutation_id)
    qoute.delete()
    return JsonResponse({'success':True, 'message':'Qoutation successfully deleted'}, status=200)

@login_required
def cashbook_view(request):
    filter_option = request.GET.get('filter', 'today')
    now = datetime.datetime.now()
    end_date = now
    
    if filter_option == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_week':
        start_date = now - timedelta(days=now.weekday())
    elif filter_option == 'yesterday':
        start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_month':
        start_date = now.replace(day=1)
    elif filter_option == 'last_month':
        start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    elif filter_option == 'this_year':
        start_date = now.replace(month=1, day=1)
    elif filter_option == 'custom':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    else:
        start_date = now - timedelta(days=now.weekday())
        end_date = now

    entries = Cashbook.objects.filter(issue_date__gte=start_date, issue_date__lte=end_date).order_by('issue_date')
    
    total_debit = entries.filter(debit=True, cancelled=False).aggregate(Sum('amount'))['amount__sum'] or 0
    total_credit = entries.filter(credit=True, cancelled=False).aggregate(Sum('amount'))['amount__sum'] or 0
    
    balance_bf = 0 
    
    previous_entries = Cashbook.objects.filter(issue_date__lt=start_date)
    previous_debit = previous_entries.filter(debit=True).aggregate(Sum('amount'))['amount__sum'] or 0
    previous_credit = previous_entries.filter(credit=True).aggregate(Sum('amount'))['amount__sum'] or 0
    balance_bf = previous_debit - previous_credit

    total_balance = total_debit - total_credit
    logger.info(total_balance)
    invoice_items = InvoiceItem.objects.all()

    return render(request, 'finance/cashbook.html', {
        'filter_option': filter_option,
        'entries': entries,
        'balance_bf': balance_bf,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'total_balance': total_balance,
        'end_date': end_date,
        'start_date': start_date,
        'invoice_items': invoice_items
    })

@login_required
def download_cashbook_report(request):
    filter_option = request.GET.get('filter', 'this_week')
    now = datetime.datetime.now()
    end_date = now
    
    if filter_option == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_week':
        start_date = now - timedelta(days=now.weekday())
    elif filter_option == 'yesterday':
        start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    elif filter_option == 'this_month':
        start_date = now.replace(day=1)
    elif filter_option == 'last_month':
        start_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    elif filter_option == 'this_year':
        start_date = now.replace(month=1, day=1)
    elif filter_option == 'custom':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
    else:
        start_date = now - timedelta(days=now.weekday())
        end_date = now

    entries = Cashbook.objects.filter(date__gte=start_date, date__lte=end_date).order_by('date')

    # Create a CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="cashbook_report_{filter_option}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Description', 'Expenses', 'Income', 'Balance'])

    balance = 0  
    for entry in entries:
        if entry.debit:
            balance += entry.amount
        elif entry.credit:
            balance -= entry.amount

        writer.writerow([
            entry.issue_date,
            entry.description,
            entry.amount if entry.debit else '',
            entry.amount if entry.credit else '',
            balance,
            entry.accountant,
            entry.manager,
            entry.director
        ])

    return response


@login_required
def cashbook_note(request):
    #payload
    """
        entry_id:id,
        note:str
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            entry_id = data.get('entry_id')
            note = data.get('note')
            
            entry = Cashbook.objects.get(id=entry_id)
            entry.note = note
            
            entry.save()
        except Exception as e:
            return JsonResponse({'success':False, 'message':f'{e}.'}, status=400)
        return JsonResponse({'success':False, 'message':'Note successfully saved.'}, status=201)
    return JsonResponse({'success':False, 'message':'Invalid request.'}, status=405)


@login_required
def cashbook_note_view(request, entry_id):
    entry = get_object_or_404(Cashbook, id=entry_id)
    
    if request.method == 'GET':
        notes = entry.notes.all().order_by('timestamp')
        notes_data = [
            {'user': note.user.username, 'note': note.note, 'timestamp': note.timestamp.strftime("%Y-%m-%d %H:%M:%S")}
            for note in notes
        ]
        return JsonResponse({'success': True, 'notes': notes_data})
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            note_text = data.get('note')
            CashBookNote.objects.create(entry=entry, user=request.user, note=note_text)
            return JsonResponse({'success': True, 'message': 'Note successfully added.'}, status=201)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request.'}, status=405)
    
@login_required
def cancel_transaction(request):
    #payload
    """
        entry_id:id,
    """
    try:
        data = json.loads(request.body)
        entry_id = int(data.get('entry_id'))
        
        logger.info(entry_id)
        
        entry = Cashbook.objects.get(id=entry_id)
        
        entry.cancelled = True
        
        if entry.director:
            entry.director = False
        elif entry.manager:
            entry.manager = False
        elif entry.accountant:
            entry.accountant = False
            
        entry.save()
        logger.info(entry)
        return JsonResponse({'success': True}, status=201)
    except Exception as e:
        logger.info(e)
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
def update_transaction_status(request, pk):
    if request.method == 'POST':
        entry = get_object_or_404(Cashbook, pk=pk)
        
        data = json.loads(request.body)
        
        status = data.get('status')
        field = data.get('field')  

        if field in ['manager', 'accountant', 'director']:
            setattr(entry, field, status)

            if entry.cancelled:
                entry.cancelled = False
            entry.save()
            return JsonResponse({'success': True, 'status': getattr(entry, field)})
        
    return JsonResponse({'success': False}, status=400)   
    
@login_required
def cashWithdrawals(request):
    search_query = request.GET.get('q', '')
    selected_query = request.GET.get('sq', '')
    
    withdrawals = CashWithdraw.objects.all().order_by('-date')
    
    if search_query:
        withdrawals = withdrawals.filter(
            Q(user__branch__name__icontains=search_query)|
            Q(amount__icontains=search_query)|
            Q(date__icontains=search_query)|
            Q(reason__icontains=search_query)
        )
    if selected_query:
        withdrawals = CashWithdraw.objects.filter(deleted=True).order_by('-date')
        
    if 'download' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=withdrawals.xlsx'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for col_num, header_title in enumerate(['Date', 'User', 'Amount', 'Reason', 'Status'], start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.alignment = header_alignment
            
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(len(header_title), 20)

        withdrawals = CashWithdraw.objects.all().order_by('-date')
        for withdrawal in withdrawals:
            worksheet.append(
                [
                    withdrawal.date,
                    withdrawal.user.username,
                    withdrawal.amount,
                    withdrawal.reason,
                    'Canceled' if withdrawal.deleted else 'Expensed' if withdrawal.status else 'pending'
                ])  
            
        workbook.save(response)
        return response
    
    form = CashWithdrawForm()
    expense_form = cashWithdrawExpenseForm()
    
    if request.method == 'POST':
        form = CashWithdrawForm(request.POST)
        
        if form.is_valid():
             
            password = form.cleaned_data['password']
            currency = form.cleaned_data['currency']
            amount = form.cleaned_data['amount']
            
            user = authenticate(username=request.user.username, password=password)
            
            if user is None:
                messages.warning(request, 'Incorrect password')
                return redirect('finance:withdrawals')
            
            cw_obj = form.save(commit=False)
            cw_obj.user = user
            cw_obj.save()
            
            account_name = f"{request.user.branch} {currency.name} {'Cash'} Account"
            
            try:
                account = Account.objects.get(name=account_name, type=Account.AccountType.CASH)
            except Account.DoesNotExist:
                messages.error(request, f'{account_name} doesnt exists')
                return redirect('finance:withdrawals')

            try:
                account_balance = AccountBalance.objects.get(account=account,  branch=request.user.branch)
            except AccountBalance.DoesNotExist:
                messages.error(request, f'Account Balances for account {account_name} doesnt exists')
                return redirect('finance:withdrawals')
            
            account_balance.balance -= Decimal(amount)
            account_balance.save()
            messages.success(request, 'Cash Withdrawal Successfully saved')
        else:
            messages.error(request, 'Invalid form data')
    return render(request, 'finance/cashWithdaraws/withdrawals.html', 
        {
            'withdrawals':withdrawals,
            'count': withdrawals.filter(status=False, deleted=False).count(),
            'expense_form':expense_form,
            'form':form,
        }
    )

@login_required
@transaction.atomic
def cash_withdrawal_to_expense(request):
    if request.method == 'GET':
        cwte_id = request.GET.get('id', '')
        withdrawals = CashWithdraw.objects.filter(id=cwte_id).values(
            'user__branch__name', 'amount', 'reason', 'currency__id', 'user__id'
        )
        return JsonResponse(list(withdrawals), safe=False)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        
        withdrawal_data = data['withdrawal'][0]
        
        reason = data['reason']
        category_id = data['category_id']
        withdrawal_id = data['withdrawal_id']
        currency_id = withdrawal_data['currency__id']
        branch_name = withdrawal_data['user__branch__name']
        amount = withdrawal_data['amount']
        
        try:
            currency = Currency.objects.get(id=currency_id)
            branch = Branch.objects.get(name=branch_name)
            withdrawal = CashWithdraw.objects.get(id=withdrawal_id)
            category = ExpenseCategory.objects.get(id=category_id)
        except:
            return JsonResponse({'success':False,'message':'Invalid form data here'})
        
        Expense.objects.create(
            category=category,
            amount=amount,
            branch=branch,
            user = request.user,
            currency = currency,
            description=reason,
            status=True,
            issue_date=withdrawal.date,
            payment_method='cash'
        )
        
        withdrawal.status=True
        withdrawal.save()
        
        return JsonResponse({'success':True, 'message':'Successfully added to expenses'}, status=201)
    return JsonResponse({'success':False, 'message':'Invalid form data'}, status=400)
       
@login_required
def delete_withdrawal(request, withdrawal_id):
    try:
        withdrawal = CashWithdraw.objects.get(id=withdrawal_id)
    except User.DoesNotExist:
        messages.warning(request, 'Withdrawal doesnt exist')
        return redirect('finance:withdrawals')
    
    account_name = f"{request.user.branch} {withdrawal.currency.name} {'Cash'} Account"
    
    try:
        account = Account.objects.get(name=account_name, type=Account.AccountType.CASH)
    except Account.DoesNotExist:
        messages.error(request, f'{account_name} doesnt exists')
        return redirect('finance:withdrawals')
    
    try:
        account_balance = AccountBalance.objects.get(account=account,  branch=request.user.branch)
    except AccountBalance.DoesNotExist:
        messages.error(request, f'Account Balances for account {account_name} doesnt exists')
        return redirect('finance:withdrawals')
    
    account_balance.balance += Decimal(withdrawal.amount)
    account_balance.save()
    withdrawal.deleted=True
    withdrawal.save()
    
    messages.success(request, 'Withdrawal successfully deleted')
    return redirect('finance:withdrawals')
    
    
@login_required
def days_data(request):
    current_month = get_current_month()

    sales = Sale.objects.filter(date__month=current_month)
    cogs = COGSItems.objects.filter(date__month=current_month)

    first_day = min(sales.first().date, cogs.first().date)
    
    def get_week_data(queryset, start_date, end_date, amount_field):
        week_data = queryset.filter(date__gte=start_date, date__lt=end_date).values(amount_field, 'date')
        logger.info(week_data)
        total = week_data.aggregate(total=Sum(amount_field))['total'] or 0
        return week_data, total

    data = {}
    for week in range(1, 5):
        week_start = first_day + timedelta(days=(week-1)*7)
        week_end = week_start + timedelta(days=7)

        logger.info(week_start)
        logger.info(week_end)

        sales_data, sales_total = get_week_data(sales, week_start, week_end, 'total_amount')
        cogs_data, cogs_total = get_week_data(cogs, week_start, week_end, 'product__cost')
        
        data[f'week {week}'] = {
            'sales': list(sales_data),
            'cogs': list(cogs_data),
            'total_sales': sales_total,
            'total_cogs': cogs_total
        }

    return JsonResponse(data)

@login_required
def income_json(request):
    current_month = get_current_month()
    today = datetime.date.today()
    
    month = request.GET.get('month', current_month)
    day = request.GET.get('day', today.day)
    
    if request.GET.get('filter') == 'today':
        sales_total = Sale.objects.filter(date=today).aggregate(Sum('total_amount'))
    else:
        sales_total = Sale.objects.filter(date__month=month).aggregate(Sum('total_amount'))

    return JsonResponse({'sales_total': sales_total['total_amount__sum'] or 0})


@login_required
def expense_json(request):
    current_month = get_current_month()
    today = datetime.date.today()
    
    month = request.GET.get('month', current_month)
    day = request.GET.get('day', today.day)
    
    if request.GET.get('filter') == 'today':
        expense_total = Expense.objects.filter(issue_date=today, status=False).aggregate(Sum('amount'))
    else:
        expense_total = Expense.objects.filter(issue_date__month=month, status=False).aggregate(Sum('amount'))
    
    return JsonResponse({'expense_total': expense_total['amount__sum'] or 0})


@login_required
def pl_overview(request):
    filter_option = request.GET.get('filter')
    today = datetime.date.today()
    previous_month = get_previous_month()
    current_year = today.year
    current_month = today.month

    if filter_option == 'today':
        date_filter = today
    elif filter_option == 'last_week':
        last_week_start = today - datetime.timedelta(days=today.weekday() + 7)
        last_week_end = last_week_start + datetime.timedelta(days=6)
        date_filter = (last_week_start, last_week_end)
    elif filter_option == 'this_month':
        date_filter = (datetime.date(current_year, current_month, 1), today)
    elif filter_option == 'year':
        year = int(request.GET.get('year', current_year))
        date_filter = (datetime.date(year, 1, 1), datetime.date(year, 12, 31))
    else:
        date_filter = (datetime.date(current_year, current_month, 1), today)

    if filter_option == 'today':
        current_month_sales = Sale.objects.filter(date=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = Expense.objects.filter(issue_date=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = COGSItems.objects.filter(date=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    elif filter_option == 'last_week':
        current_month_sales = Sale.objects.filter(date__range=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = Expense.objects.filter(issue_date__range=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = COGSItems.objects.filter(date__range=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    else:
        current_month_sales = Sale.objects.filter(date__range=date_filter).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        current_month_expenses = Expense.objects.filter(dissue_date__range=date_filter).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
        cogs_total = COGSItems.objects.filter(date__range=date_filter).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0

    previous_month_sales = Sale.objects.filter(date__year=current_year, date__month=previous_month).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
    previous_month_expenses = Expense.objects.filter(issue_date__year=current_year, issue_date__month=previous_month).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0
    previous_cogs =  COGSItems.objects.filter(date__year=current_year, date__month=previous_month).aggregate(total_cogs=Sum('product__cost'))['total_cogs'] or 0
    
    current_net_income = current_month_sales
    previous_net_income = previous_month_sales 
    current_expenses = current_month_expenses 
    
    current_gross_profit = current_month_sales - cogs_total
    previous_gross_profit = previous_month_sales - previous_cogs
    
    current_net_profit = current_gross_profit - current_month_expenses
    previous_net_profit = previous_gross_profit - previous_month_expenses

    current_gross_profit_margin = (current_gross_profit / current_month_sales * 100) if current_month_sales != 0 else 0
    previous_gross_profit_margin = (previous_gross_profit / previous_month_sales * 100) if previous_month_sales != 0 else 0
    
    # net_income_change = calculate_percentage_change(current_net_income, previous_net_income)
    # gross_profit_change = calculate_percentage_change(current_gross_profit, previous_gross_profit)
    # gross_profit_margin_change = calculate_percentage_change(current_gross_profit_margin, previous_gross_profit_margin)


    data = {
        'net_profit':current_net_profit,
        'cogs_total':cogs_total,
        'current_expenses':current_expenses,
        'current_net_profit': current_net_profit,
        'previous_net_profit':previous_net_profit,
        'current_net_income': current_net_income,
        'previous_net_income': previous_net_income,
        'current_gross_profit': current_gross_profit,
        'previous_gross_profit': previous_gross_profit,
        'current_gross_profit_margin': f'{current_gross_profit_margin:.2f}',
        'previous_gross_profit_margin': previous_gross_profit_margin,
    }
    
    return JsonResponse(data)
